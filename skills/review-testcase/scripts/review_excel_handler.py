#!/usr/bin/env python3
"""
Excel测试用例评审处理器
用于读取Excel测试用例文件，追加评审结果列和评审汇总Sheet

关键设计：
- 在原始用例列后新增9列评审字段（分类、总分、5维得分、扣分原因、改进建议）
- 用例分类列自动设置颜色标记（可用=绿色，待修改=黄色，错误无效=红色）
- 新增"评审汇总"Sheet，包含总体概况、指标达标情况、维度得分分布、典型问题
- 保留原始用例内容和格式
"""

import sys
import os
import copy
import json
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# 评审新增列定义
REVIEW_COLUMNS = [
    '用例分类',          # 可用/待修改/错误无效
    '评审总分',          # 0-100
    '逻辑完整性',        # 0-25
    '预期结果明确性',    # 0-20
    '前置条件完备性',    # 0-15
    'PRD覆盖度',         # 0-25
    '边界异常覆盖',      # 0-15
    '扣分原因',          # 具体扣分项
    '改进建议',          # 具体改进措施
]

# 分类颜色映射
CLASSIFICATION_COLORS = {
    '可用': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),       # 绿色
    '待修改': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),     # 黄色
    '错误无效': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),   # 红色
}

# 样式定义
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
NORMAL_FONT = Font(size=10)
BOLD_FONT = Font(size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical='top')


def read_testcase_excel(file_path):
    """
    读取Excel测试用例文件，返回结构化数据

    Args:
        file_path: Excel文件路径

    Returns:
        dict: 包含文件、工作簿、Sheet信息的结构化数据
    """
    if not os.path.exists(file_path):
        print(f"ERROR: File not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(file_path)
    result = {
        'file_path': file_path,
        'wb': wb,
        'sheets': []
    }

    for ws in wb.worksheets:
        sheet_data = {
            'name': ws.title,
            'headers': [],
            'rows': [],
            'max_row': ws.max_row,
            'header_row': 1,
            'ws': ws
        }

        # 查找标题行
        header_row_idx = 1
        for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=False):
            non_empty = [cell.value for cell in row if cell.value is not None]
            if len(non_empty) >= 3:
                header_row_idx = row[0].row
                break

        sheet_data['header_row'] = header_row_idx

        # 读取标题
        for cell in ws[header_row_idx]:
            header_val = cell.value if cell.value is not None else f"Column_{cell.column}"
            sheet_data['headers'].append(str(header_val).strip())

        # 读取数据行
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(sheet_data['headers'], 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data[header] = cell.value
            if any(v is not None for v in row_data.values()):
                row_data['_row_idx'] = row_idx
                sheet_data['rows'].append(row_data)

        result['sheets'].append(sheet_data)

    return result


def find_column_index(headers, keywords):
    """
    在标题列表中查找包含关键词的列索引

    Args:
        headers: 标题列表
        keywords: 关键词列表

    Returns:
        int: 列索引（0-based），未找到返回-1
    """
    for idx, header in enumerate(headers):
        header_lower = header.lower()
        for kw in keywords:
            if kw in header_lower:
                return idx
    return -1


def append_review_to_excel(file_path, review_results, metrics_data, output_path=None):
    """
    向Excel文件追加评审结果

    Args:
        file_path: 原始Excel文件路径
        review_results: 评审结果列表，每条用例的评审数据，格式：
            [
                {
                    'row_index': 行号,
                    'classification': '可用/待修改/错误无效',
                    'total_score': 85,
                    'scores': {
                        '逻辑完整性': 22,
                        '预期结果明确性': 18,
                        '前置条件完备性': 13,
                        'PRD覆盖度': 20,
                        '边界异常覆盖': 12
                    },
                    'deduction_reasons': '步骤跳跃-3分 | 预期结果模糊-2分',
                    'improvement_suggestions': '1. 补充第3步中间操作\n2. 明确第2步预期结果'
                },
                ...
            ]
        metrics_data: 量化指标数据，格式：
            {
                'total_cases': 100,
                'classification_summary': {'可用': 60, '待修改': 30, '错误无效': 10},
                'metrics': {
                    '需求覆盖率': {'value': '96%', 'threshold': '≥ 95%', 'passed': True},
                    ...
                },
                'dimension_scores': {
                    '逻辑完整性': {'avg': 20.5, 'max': 25, 'rate': '82%'},
                    ...
                },
                'top_issues': [
                    {'rank': 1, 'type': '预期结果模糊', 'count': 25, 'example': '"验证成功"无具体判定标准'},
                    ...
                ]
            }
        output_path: 输出文件路径

    Returns:
        str: 输出文件路径
    """
    data = read_testcase_excel(file_path)

    if not output_path:
        base, ext = os.path.splitext(file_path)
        output_path = f"{base}【评审版】{ext}"

    wb = data['wb']
    sheet_data = data['sheets'][0]
    ws = sheet_data['ws']
    headers = list(sheet_data['headers'])

    # 新增评审列
    review_col_start = len(headers) + 1
    for i, col_name in enumerate(REVIEW_COLUMNS):
        col_idx = review_col_start + i
        cell = ws.cell(row=sheet_data['header_row'], column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
        headers.append(col_name)

    # 写入评审数据
    for result in review_results:
        row_idx = result.get('row_index', 0)
        if row_idx <= 0:
            continue

        classification = result.get('classification', '')
        total_score = result.get('total_score', 0)
        scores = result.get('scores', {})
        deduction_reasons = result.get('deduction_reasons', '')
        suggestions = result.get('improvement_suggestions', '')

        # 写入分类列
        col_offset = 0
        class_cell = ws.cell(row=row_idx, column=review_col_start + col_offset, value=classification)
        if classification in CLASSIFICATION_COLORS:
            class_cell.fill = CLASSIFICATION_COLORS[classification]
        class_cell.alignment = Alignment(horizontal='center', vertical='top')
        class_cell.border = THIN_BORDER

        # 写入总分
        col_offset = 1
        score_cell = ws.cell(row=row_idx, column=review_col_start + col_offset, value=total_score)
        score_cell.alignment = Alignment(horizontal='center', vertical='top')
        score_cell.border = THIN_BORDER

        # 写入5维得分
        score_keys = ['逻辑完整性', '预期结果明确性', '前置条件完备性', 'PRD覆盖度', '边界异常覆盖']
        for j, key in enumerate(score_keys):
            col_offset = 2 + j
            dim_cell = ws.cell(row=row_idx, column=review_col_start + col_offset, value=scores.get(key, 0))
            dim_cell.alignment = Alignment(horizontal='center', vertical='top')
            dim_cell.border = THIN_BORDER

        # 写入扣分原因
        col_offset = 7
        reason_cell = ws.cell(row=row_idx, column=review_col_start + col_offset, value=deduction_reasons)
        reason_cell.alignment = WRAP_ALIGNMENT
        reason_cell.border = THIN_BORDER

        # 写入改进建议
        col_offset = 8
        suggest_cell = ws.cell(row=row_idx, column=review_col_start + col_offset, value=suggestions)
        suggest_cell.alignment = WRAP_ALIGNMENT
        suggest_cell.border = THIN_BORDER

    # 设置评审列宽度
    col_widths = {
        '用例分类': 12, '评审总分': 10, '逻辑完整性': 12, '预期结果明确性': 14,
        '前置条件完备性': 14, 'PRD覆盖度': 12, '边界异常覆盖': 12,
        '扣分原因': 40, '改进建议': 40
    }
    for i, col_name in enumerate(REVIEW_COLUMNS):
        col_idx = review_col_start + i
        width = col_widths.get(col_name, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 新增"评审汇总"Sheet
    _create_summary_sheet(wb, metrics_data)

    wb.save(output_path)
    print(f"Review saved to: {output_path}")
    return output_path


def _create_summary_sheet(wb, metrics_data):
    """
    创建评审汇总Sheet

    Args:
        wb: workbook对象
        metrics_data: 量化指标数据
    """
    ws_name = '评审汇总'
    # 如果已存在同名Sheet，先删除
    if ws_name in wb.sheetnames:
        del wb[ws_name]

    ws = wb.create_sheet(title=ws_name)

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40

    current_row = 1

    # 标题
    ws.cell(row=current_row, column=1, value='AI 测试用例评审汇总报告')
    ws.cell(row=current_row, column=1).font = Font(size=16, bold=True)
    current_row += 2

    # 一、总体概况
    ws.cell(row=current_row, column=1, value='一、总体概况')
    ws.cell(row=current_row, column=1).font = Font(size=13, bold=True)
    current_row += 1

    total = metrics_data.get('total_cases', 0)
    summary = metrics_data.get('classification_summary', {})
    for cls_name, count in summary.items():
        pct = f"{count / total * 100:.1f}%" if total > 0 else '0%'
        label = f"{cls_name}（{'第一类' if cls_name == '可用' else '第二类' if cls_name == '待修改' else '第三类'}）"
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=f"{count}（{pct}）")
        if cls_name in CLASSIFICATION_COLORS:
            ws.cell(row=current_row, column=1).fill = CLASSIFICATION_COLORS[cls_name]
        current_row += 1

    ws.cell(row=current_row, column=1, value='总用例数')
    ws.cell(row=current_row, column=2, value=total)
    current_row += 2

    # 二、量化指标达标情况
    ws.cell(row=current_row, column=1, value='二、量化指标达标情况')
    ws.cell(row=current_row, column=1).font = Font(size=13, bold=True)
    current_row += 1

    # 表头
    for col, header in enumerate(['指标', '实际值', '合格线', '是否达标', '备注'], 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    current_row += 1

    metrics = metrics_data.get('metrics', {})
    for metric_name, metric_info in metrics.items():
        value = metric_info.get('value', '-')
        threshold = metric_info.get('threshold', '-')
        passed = metric_info.get('passed', None)
        note = metric_info.get('note', '')

        ws.cell(row=current_row, column=1, value=metric_name).border = THIN_BORDER
        ws.cell(row=current_row, column=2, value=value).border = THIN_BORDER
        ws.cell(row=current_row, column=3, value=threshold).border = THIN_BORDER

        status_cell = ws.cell(row=current_row, column=4)
        if passed is True:
            status_cell.value = '✅ 达标'
            status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif passed is False:
            status_cell.value = '❌ 不达标'
            status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        else:
            status_cell.value = '⚠️ 无法计算'
            status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        status_cell.border = THIN_BORDER

        ws.cell(row=current_row, column=5, value=note).border = THIN_BORDER
        current_row += 1

    current_row += 1

    # 三、各维度得分分布
    ws.cell(row=current_row, column=1, value='三、各维度得分分布')
    ws.cell(row=current_row, column=1).font = Font(size=13, bold=True)
    current_row += 1

    for col, header in enumerate(['维度', '平均分', '满分', '得分率'], 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    current_row += 1

    dim_scores = metrics_data.get('dimension_scores', {})
    for dim_name, dim_info in dim_scores.items():
        avg = dim_info.get('avg', 0)
        max_score = dim_info.get('max', 0)
        rate = dim_info.get('rate', '0%')

        ws.cell(row=current_row, column=1, value=dim_name).border = THIN_BORDER
        ws.cell(row=current_row, column=2, value=avg).border = THIN_BORDER
        ws.cell(row=current_row, column=3, value=max_score).border = THIN_BORDER
        ws.cell(row=current_row, column=4, value=rate).border = THIN_BORDER
        current_row += 1

    current_row += 1

    # 四、典型问题 Top 10
    ws.cell(row=current_row, column=1, value='四、典型问题 Top 10')
    ws.cell(row=current_row, column=1).font = Font(size=13, bold=True)
    current_row += 1

    for col, header in enumerate(['排名', '问题类型', '出现次数', '典型示例'], 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    current_row += 1

    top_issues = metrics_data.get('top_issues', [])
    for issue in top_issues[:10]:
        ws.cell(row=current_row, column=1, value=issue.get('rank', '')).border = THIN_BORDER
        ws.cell(row=current_row, column=2, value=issue.get('type', '')).border = THIN_BORDER
        ws.cell(row=current_row, column=3, value=issue.get('count', '')).border = THIN_BORDER
        ws.cell(row=current_row, column=4, value=issue.get('example', '')).border = THIN_BORDER
        current_row += 1


def print_sheet_summary(data):
    """打印Excel文件概要信息"""
    print(f"文件: {data['file_path']}")
    for sheet in data['sheets']:
        print(f"\nSheet: {sheet['name']}")
        print(f"  标题行: 第{sheet['header_row']}行")
        print(f"  列标题: {sheet['headers']}")
        print(f"  数据行数: {len(sheet['rows'])}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage:")
        print(f"  python {sys.argv[0]} read <excel_file>          - 读取并打印Excel用例概要")
        print(f"  python {sys.argv[0]} review <excel_file>         - 追加示例评审结果（测试用）")
        sys.exit(0)

    command = sys.argv[1]

    if command == 'read':
        if len(sys.argv) < 3:
            print("ERROR: Please provide Excel file path", file=sys.stderr)
            sys.exit(1)
        data = read_testcase_excel(sys.argv[2])
        print_sheet_summary(data)

    elif command == 'review':
        if len(sys.argv) < 3:
            print("ERROR: Please provide Excel file path", file=sys.stderr)
            sys.exit(1)
        # 测试评审追加功能
        test_review = [
            {
                'row_index': 2,
                'classification': '可用',
                'total_score': 88,
                'scores': {
                    '逻辑完整性': 22,
                    '预期结果明确性': 18,
                    '前置条件完备性': 13,
                    'PRD覆盖度': 20,
                    '边界异常覆盖': 15
                },
                'deduction_reasons': '步骤跳跃-3分',
                'improvement_suggestions': '补充中间操作步骤'
            },
            {
                'row_index': 3,
                'classification': '待修改',
                'total_score': 65,
                'scores': {
                    '逻辑完整性': 18,
                    '预期结果明确性': 12,
                    '前置条件完备性': 10,
                    'PRD覆盖度': 15,
                    '边界异常覆盖': 10
                },
                'deduction_reasons': '预期结果模糊-8分 | 前置条件缺失-5分',
                'improvement_suggestions': '1. 明确预期结果的具体判定标准\n2. 补充账号权限和订单状态'
            },
            {
                'row_index': 4,
                'classification': '错误无效',
                'total_score': 30,
                'scores': {
                    '逻辑完整性': 8,
                    '预期结果明确性': 5,
                    '前置条件完备性': 3,
                    'PRD覆盖度': 8,
                    '边界异常覆盖': 6
                },
                'deduction_reasons': '业务逻辑错误-10分 | 步骤自相矛盾-5分 | 预期结果与规则不符',
                'improvement_suggestions': '废弃此用例，重新设计'
            }
        ]
        test_metrics = {
            'total_cases': 3,
            'classification_summary': {'可用': 1, '待修改': 1, '错误无效': 1},
            'metrics': {
                '反向用例占比': {'value': '33%', 'threshold': '≥ 30%', 'passed': True},
                '错误用例率': {'value': '33%', 'threshold': '≤ 5%', 'passed': False},
                '重复用例率': {'value': '0%', 'threshold': '≤ 10%', 'passed': True},
            },
            'dimension_scores': {
                '逻辑完整性': {'avg': 16.0, 'max': 25, 'rate': '64%'},
                '预期结果明确性': {'avg': 11.7, 'max': 20, 'rate': '58%'},
                '前置条件完备性': {'avg': 8.7, 'max': 15, 'rate': '58%'},
                'PRD覆盖度': {'avg': 14.3, 'max': 25, 'rate': '57%'},
                '边界异常覆盖': {'avg': 10.3, 'max': 15, 'rate': '69%'},
            },
            'top_issues': [
                {'rank': 1, 'type': '预期结果模糊', 'count': 2, 'example': '"验证成功"无具体判定标准'},
                {'rank': 2, 'type': '前置条件缺失', 'count': 1, 'example': '未说明账号权限'},
            ]
        }
        output = append_review_to_excel(sys.argv[2], test_review, test_metrics)
        print(f"Output saved to: {output}")

    else:
        print(f"Unknown command: {command}", file=sys.stderr)
        sys.exit(1)
