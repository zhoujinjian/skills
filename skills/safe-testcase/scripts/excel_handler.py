#!/usr/bin/env python3
"""
Excel测试用例文件处理器
用于读取测试用例Excel文件，追加补全用例，生成场景补全版文件

关键设计：
- 补全标识放在独立列"补充标识"，不在用例标题中添加标记
- 原始用例的"补充标识"列留空，补全用例填写"AI补全-场景遗漏"
- 便于Excel中按"补充标识"列筛选过滤，不破坏用例标题原有语义
"""

import sys
import os
import copy
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# 补充标识列的默认列名
SUPPLEMENT_TAG_COLUMN = '补充标识'
# AI补全标识值
AI_SUPPLEMENT_TAG = 'AI补全-场景遗漏'


def read_testcase_excel(file_path):
    """
    读取Excel测试用例文件，返回结构化数据

    Args:
        file_path: Excel文件路径

    Returns:
        dict: {
            'file_path': 文件路径,
            'wb': workbook对象,
            'sheets': [
                {
                    'name': sheet名称,
                    'headers': 列标题列表,
                    'rows': 行数据列表(每行为字典),
                    'max_row': 最大数据行号,
                    'header_row': 标题行号
                }
            ]
        }
    """
    if not os.path.exists(file_path):
        print(f"ERROR: File not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(file_path)
    result = {
        'file_path': file_path,
        'wb': wb,
        'sheets': []
    }

    for ws in wb.worksheets:
        sheet_data = {
            'name': ws.title,
            'headers': [],
            'rows': [],
            'max_row': ws.max_row,
            'header_row': 1,
            'ws': ws
        }

        # 查找标题行（通常第一行，但也可能是合并单元格的情况）
        header_row_idx = 1
        for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=False):
            non_empty = [cell.value for cell in row if cell.value is not None]
            if len(non_empty) >= 3:  # 至少3列有值才认为是标题行
                header_row_idx = row[0].row
                break

        sheet_data['header_row'] = header_row_idx

        # 读取标题
        for cell in ws[header_row_idx]:
            header_val = cell.value if cell.value is not None else f"Column_{cell.column}"
            sheet_data['headers'].append(str(header_val).strip())

        # 读取数据行
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(sheet_data['headers'], 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data[header] = cell.value
            # 跳过全空行
            if any(v is not None for v in row_data.values()):
                row_data['_row_idx'] = row_idx
                sheet_data['rows'].append(row_data)

        result['sheets'].append(sheet_data)

    return result


def find_column_index(headers, keywords):
    """
    在标题列表中查找包含关键词的列索引

    Args:
        headers: 标题列表
        keywords: 关键词列表，任一匹配即可

    Returns:
        int: 列索引（0-based），未找到返回-1
    """
    for idx, header in enumerate(headers):
        header_lower = header.lower()
        for kw in keywords:
            if kw in header_lower:
                return idx
    return -1


def get_max_case_number(sheet_data):
    """
    获取当前sheet中最大的用例编号

    Args:
        sheet_data: sheet数据字典

    Returns:
        int: 最大用例编号，无编号返回0
    """
    headers = sheet_data['headers']
    # 尝试查找编号列
    col_idx = find_column_index(headers, ['编号', '序号', '用例编号', 'id', 'case_id', 'no'])
    if col_idx == -1:
        return 0

    max_num = 0
    for row in sheet_data['rows']:
        val = row.get(headers[col_idx])
        if val is not None:
            try:
                # 处理如 TC001, CASE-001 等格式
                num_str = str(val)
                numeric_part = ''
                for ch in reversed(num_str):
                    if ch.isdigit():
                        numeric_part = ch + numeric_part
                    elif numeric_part:
                        break
                if numeric_part:
                    max_num = max(max_num, int(numeric_part))
                else:
                    max_num = max(max_num, int(float(val)))
            except (ValueError, TypeError):
                continue

    return max_num


def generate_case_number(max_num, index, prefix=''):
    """
    生成用例编号

    Args:
        max_num: 当前最大编号
        index: 补充用例的序号（从1开始）
        prefix: 编号前缀（如TC、CASE-等）

    Returns:
        str: 生成的用例编号
    """
    num = max_num + index
    if prefix:
        return f"{prefix}{num:03d}"
    return str(num)


def _ensure_supplement_column(ws, headers, header_row, max_row):
    """
    确保Excel中存在"补充标识"列

    如果原始Excel中没有"补充标识"列，则在最后一列后新增该列：
    - 在标题行添加列名
    - 原始数据行的该列留空（表示非AI补全）
    - 复制相邻列的样式

    Args:
        ws: worksheet对象
        headers: 列标题列表（会被修改）
        header_row: 标题行号
        max_row: 最大数据行号

    Returns:
        int: "补充标识"列的1-based索引
    """
    # 检查是否已有补充标识列
    tag_col_idx = find_column_index(headers, ['补充标识', '补全标识', '来源标识', 'supplement', 'tag'])
    if tag_col_idx != -1:
        # 已存在，返回1-based列号
        return tag_col_idx + 1

    # 不存在，新增列
    new_col = len(headers) + 1  # 1-based

    # 添加标题
    ws.cell(row=header_row, column=new_col, value=SUPPLEMENT_TAG_COLUMN)

    # 复制前一列标题的样式
    prev_cell = ws.cell(row=header_row, column=new_col - 1)
    new_cell = ws.cell(row=header_row, column=new_col)
    if prev_cell.has_style:
        new_cell.font = copy.copy(prev_cell.font)
        new_cell.border = copy.copy(prev_cell.border)
        new_cell.fill = copy.copy(prev_cell.fill)
        new_cell.alignment = copy.copy(prev_cell.alignment)

    # 原始数据行留空
    for row_idx in range(header_row + 1, max_row + 1):
        cell = ws.cell(row=row_idx, column=new_col)
        cell.value = None  # 原始用例的标识列为空
        # 复制样式
        if max_row > header_row:
            src_cell = ws.cell(row=row_idx, column=new_col - 1)
            if src_cell.has_style:
                cell.font = copy.copy(src_cell.font)
                cell.border = copy.copy(src_cell.border)
                cell.fill = copy.copy(src_cell.fill)
                cell.alignment = copy.copy(src_cell.alignment)

    # 更新headers列表
    headers.append(SUPPLEMENT_TAG_COLUMN)

    return new_col


def append_testcases_to_excel(file_path, new_cases, output_path=None):
    """
    向Excel文件追加补全用例

    关键行为：
    - 自动检查并新增"补充标识"列（如原始Excel中不存在）
    - 补全用例的"补充标识"列填写"AI补全-场景遗漏"
    - 原始用例的"补充标识"列留空
    - 用例标题中不再追加标识标记，保持标题语义完整
    - new_cases中如包含"补充标识"key，使用其值；否则默认填入AI_SUPPLEMENT_TAG

    Args:
        file_path: 原始Excel文件路径
        new_cases: 新用例列表，每个用例为字典，key对应列标题
                   注意：用例标题中不再需要包含【AI补全-场景遗漏】标记
                   标识信息会自动写入"补充标识"列
        output_path: 输出文件路径，默认为 原文件名+【场景补全版】.xlsx

    Returns:
        str: 输出文件路径
    """
    data = read_testcase_excel(file_path)

    if not output_path:
        base, ext = os.path.splitext(file_path)
        output_path = f"{base}【场景补全版】{ext}"

    wb = data['wb']

    # 默认处理第一个sheet
    sheet_data = data['sheets'][0]
    ws = sheet_data['ws']
    headers = sheet_data['headers']

    # 确保"补充标识"列存在
    tag_col = _ensure_supplement_column(ws, headers, sheet_data['header_row'], sheet_data['max_row'])

    # 获取最大编号
    max_num = get_max_case_number(sheet_data)

    # 查找编号列的前缀格式
    num_col_idx = find_column_index(headers, ['编号', '序号', '用例编号', 'id', 'case_id', 'no'])
    prefix = ''
    if num_col_idx != -1 and sheet_data['rows']:
        first_val = sheet_data['rows'][0].get(headers[num_col_idx])
        if first_val is not None:
            val_str = str(first_val)
            # 提取前缀部分（非数字部分）
            for i, ch in enumerate(val_str):
                if ch.isdigit():
                    prefix = val_str[:i]
                    break

    # 追加新用例
    next_row = sheet_data['max_row'] + 1
    for case_idx, case in enumerate(new_cases, 1):
        for col_idx, header in enumerate(headers, 1):
            if header == SUPPLEMENT_TAG_COLUMN:
                # 补充标识列：填写AI补全标记
                tag_value = case.get(SUPPLEMENT_TAG_COLUMN, AI_SUPPLEMENT_TAG)
                ws.cell(row=next_row, column=col_idx, value=tag_value)
            else:
                value = case.get(header, '')
                if value is None:
                    value = ''
                ws.cell(row=next_row, column=col_idx, value=value)

            # 复制上一行的样式
            if sheet_data['max_row'] > sheet_data['header_row']:
                src_cell = ws.cell(row=sheet_data['max_row'], column=col_idx)
                dst_cell = ws.cell(row=next_row, column=col_idx)
                if src_cell.has_style:
                    dst_cell.font = copy.copy(src_cell.font)
                    dst_cell.border = copy.copy(src_cell.border)
                    dst_cell.fill = copy.copy(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.protection = copy.copy(src_cell.protection)
                    dst_cell.alignment = copy.copy(src_cell.alignment)

        # 自动填充编号列
        if num_col_idx != -1:
            case_num = generate_case_number(max_num, case_idx, prefix)
            ws.cell(row=next_row, column=num_col_idx + 1, value=case_num)

        next_row += 1

    wb.save(output_path)
    return output_path


def print_sheet_summary(data):
    """打印Excel文件概要信息"""
    print(f"文件: {data['file_path']}")
    for sheet in data['sheets']:
        print(f"\nSheet: {sheet['name']}")
        print(f"  标题行: 第{sheet['header_row']}行")
        print(f"  列标题: {sheet['headers']}")
        print(f"  数据行数: {len(sheet['rows'])}")
        max_num = get_max_case_number(sheet)
        print(f"  最大用例编号: {max_num}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage:")
        print(f"  python {sys.argv[0]} read <excel_file>          - 读取并打印Excel用例概要")
        print(f"  python {sys.argv[0]} append <excel_file>         - 追加示例用例（测试用）")
        sys.exit(0)

    command = sys.argv[1]

    if command == 'read':
        if len(sys.argv) < 3:
            print("ERROR: Please provide Excel file path", file=sys.stderr)
            sys.exit(1)
        data = read_testcase_excel(sys.argv[2])
        print_sheet_summary(data)

    elif command == 'append':
        if len(sys.argv) < 3:
            print("ERROR: Please provide Excel file path", file=sys.stderr)
            sys.exit(1)
        # 测试追加功能 - 注意：标题中不再包含标识，标识写入"补充标识"列
        test_case = {
            '用例标题': '测试补全用例',
            '前置条件': '测试环境已就绪',
            '测试步骤': '1. 执行测试操作',
            '预期结果': '系统按预期响应',
            '优先级': 'P2'
        }
        output = append_testcases_to_excel(sys.argv[2], [test_case])
        print(f"Output saved to: {output}")

    else:
        print(f"Unknown command: {command}", file=sys.stderr)
        sys.exit(1)
