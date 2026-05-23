#!/usr/bin/env python3
"""
测试用例 Excel 生成器

根据 JSON 格式的测试用例数据，生成格式化的 Excel (.xlsx) 文件。

输入: JSON 格式的测试用例数据（通过命令行参数指定 JSON 文件路径）
输出: .xlsx 文件

用法:
    python generate_excel.py <input.json> [output.xlsx]

JSON 输入格式:
{
    "title": "项目名称 - 测试用例",
    "test_cases": [
        {
            "id": "TC-001",
            "title": "用例标题",
            "precondition": "前置条件",
            "steps": "1. 步骤一\\n2. 步骤二\\n3. 步骤三",
            "test_data": "测试数据描述",
            "expected_result": "预期结果描述",
            "priority": "P0",
            "requirement_id": "US-001"
        }
    ]
}
"""

import json
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ── 样式定义 ──────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

P0_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
P1_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
P2_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

PRIORITY_FILLS = {"P0": P0_FILL, "P1": P1_FILL, "P2": P2_FILL}

CELL_FONT = Font(name="Microsoft YaHei", size=10)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 列定义: (表头名称, 列宽)
COLUMNS = [
    ("用例编号", 12),
    ("用例标题", 40),
    ("前置条件", 25),
    ("测试步骤", 45),
    ("测试数据", 30),
    ("预期结果", 40),
    ("优先级", 10),
    ("关联需求", 14),
]


def create_header(ws):
    """创建表头行"""
    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_test_case(ws, row, tc):
    """写入单个测试用例"""
    values = [
        tc.get("id", ""),
        tc.get("title", ""),
        tc.get("precondition", ""),
        tc.get("steps", ""),
        tc.get("test_data", ""),
        tc.get("expected_result", ""),
        tc.get("priority", ""),
        tc.get("requirement_id", ""),
    ]

    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER

        # 优先级列居中 + 条件填充
        if col_idx == 7:
            cell.alignment = CENTER_ALIGNMENT
            priority = str(value).upper()
            if priority in PRIORITY_FILLS:
                cell.fill = PRIORITY_FILLS[priority]
        # 编号列居中
        elif col_idx == 1:
            cell.alignment = CENTER_ALIGNMENT
        # 关联需求列居中
        elif col_idx == 8:
            cell.alignment = CENTER_ALIGNMENT
        else:
            cell.alignment = CELL_ALIGNMENT


def generate_excel(data, output_path):
    """生成 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 冻结首行
    ws.freeze_panes = "A2"

    # 设置自动筛选
    ws.auto_filter.ref = f"A1:H1"

    # 创建表头
    create_header(ws)

    # 写入测试用例
    test_cases = data.get("test_cases", [])
    for idx, tc in enumerate(test_cases, 2):
        write_test_case(ws, idx, tc)

    # 设置行高
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(test_cases) + 2):
        ws.row_dimensions[row_idx].height = 45

    # 添加统计 Sheet
    ws_stats = wb.create_sheet(title="统计")
    write_stats(ws_stats, data)

    # 保存
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    return output_path


def write_stats(ws, data):
    """写入统计 Sheet"""
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    # 标题
    ws.cell(row=1, column=1, value="测试用例统计").font = Font(
        name="Microsoft YaHei", size=14, bold=True
    )

    # 项目信息
    ws.cell(row=3, column=1, value="项目名称").font = Font(bold=True)
    ws.cell(row=3, column=2, value=data.get("title", ""))

    test_cases = data.get("test_cases", [])
    ws.cell(row=4, column=1, value="用例总数").font = Font(bold=True)
    ws.cell(row=4, column=2, value=len(test_cases))

    # 优先级统计
    ws.cell(row=6, column=1, value="优先级统计").font = Font(bold=True, size=12)

    headers = ["优先级", "数量", "占比"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    priority_counts = {"P0": 0, "P1": 0, "P2": 0}
    for tc in test_cases:
        p = tc.get("priority", "").upper()
        if p in priority_counts:
            priority_counts[p] += 1

    total = len(test_cases) or 1
    for row_idx, (p, count) in enumerate(priority_counts.items(), 8):
        ws.cell(row=row_idx, column=1, value=p).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=count).border = THIN_BORDER
        pct = f"{count / total * 100:.1f}%"
        ws.cell(row=row_idx, column=3, value=pct).border = THIN_BORDER
        if p in PRIORITY_FILLS:
            ws.cell(row=row_idx, column=1).fill = PRIORITY_FILLS[p]

    # 关联需求统计
    ws.cell(row=12, column=1, value="关联需求统计").font = Font(bold=True, size=12)

    req_counts = {}
    for tc in test_cases:
        req = tc.get("requirement_id", "")
        if req:
            req_counts[req] = req_counts.get(req, 0) + 1

    headers2 = ["需求编号", "用例数量"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws.cell(row=13, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row_idx, (req, count) in enumerate(sorted(req_counts.items()), 14):
        ws.cell(row=row_idx, column=1, value=req).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=count).border = THIN_BORDER


def main():
    if len(sys.argv) < 2:
        print("用法: python generate_excel.py <input.json> [output.xlsx]")
        print("  input.json   - 测试用例 JSON 数据文件")
        print("  output.xlsx  - 输出的 Excel 文件路径（可选，默认与输入同目录）")
        sys.exit(1)

    input_path = sys.argv[1]

    if not os.path.exists(input_path):
        print(f"错误: 输入文件不存在: {input_path}")
        sys.exit(1)

    # 确定输出路径
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        base_name = Path(input_path).stem
        output_dir = Path(input_path).parent
        output_path = str(output_dir / f"{base_name}.xlsx")

    # 读取输入数据
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # 生成 Excel 文件
    result = generate_excel(data, output_path)
    print(f"Excel 文件已生成: {result}")
    return result


if __name__ == "__main__":
    main()
