#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
API Test Data Generator - 测试数据自动化构造引擎

支持三种输入模式：
  1. 传统模式（api）：基于 api_definitions.json 接口定义生成测试数据
  2. 规范模式（spec）：基于自定义字段规则文件生成测试数据
  3. 自然语言模式（nl）：基于自然语言描述，Faker 驱动生成数据

四大数据维度：正向合法数据、边界值数据、异常非法数据、安全与幂等数据

Usage:
    # 传统模式
    python3 testdata_generator.py api_definitions.json [--format yaml|json|excel]
                                     [--output <dir>] [--module <name>]
                                     [--api <api_id>] [--config <rules.yaml>]
                                     [--dimension positive,boundary,negative,security]

    # 规范模式
    python3 testdata_generator.py --mode spec --spec <rules.yaml>
                                     [--format csv|yaml|json|excel]
                                     [--output <dir>] [--dimension ...]

    # 自然语言模式
    python3 testdata_generator.py --mode nl --prompt "生成10条手机号"
                                     [--output <dir>]
"""

import argparse
import copy
import csv
import io
import json
import os
import random
import re
import string
import sys
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional, Set, Tuple

# ============================================================================
# 常量定义
# ============================================================================

VERSION = "2.0.0"

# 默认字符串长度范围
DEFAULT_MIN_LENGTH = 1
DEFAULT_MAX_LENGTH = 255

# 默认数值范围
DEFAULT_INT_MIN = -2147483648
DEFAULT_INT_MAX = 2147483647

# 默认数组长度范围
DEFAULT_MIN_ITEMS = 0
DEFAULT_MAX_ITEMS = 100

# 安全注入载荷
SQL_INJECTION_PAYLOADS = [
    {"id": "SQL-001", "payload": "' OR '1'='1", "description": "SQL注入-经典OR绕过"},
    {"id": "SQL-002", "payload": "' OR '1'='1' --", "description": "SQL注入-注释截断"},
    {"id": "SQL-003", "payload": "1' UNION SELECT * FROM users--", "description": "SQL注入-联合查询"},
    {"id": "SQL-004", "payload": "'; DROP TABLE users;--", "description": "SQL注入-删表攻击"},
]

XSS_PAYLOADS = [
    {"id": "XSS-001", "payload": "<script>alert('XSS')</script>", "description": "XSS-script标签"},
    {"id": "XSS-002", "payload": "<img src=x onerror=alert(1)>", "description": "XSS-img标签"},
    {"id": "XSS-003", "payload": "<svg/onload=alert(1)>", "description": "XSS-SVG标签"},
    {"id": "XSS-004", "payload": "javascript:alert(1)", "description": "XSS-javascript协议"},
]

PATH_TRAVERSAL_PAYLOADS = [
    {"id": "PATH-001", "payload": "../../etc/passwd", "description": "路径穿越-Linux"},
    {"id": "PATH-002", "payload": "..\\..\\windows\\system32", "description": "路径穿越-Windows"},
]

COMMAND_INJECTION_PAYLOADS = [
    {"id": "CMD-001", "payload": "; ls -la", "description": "命令注入-分号"},
    {"id": "CMD-002", "payload": "$(whoami)", "description": "命令注入-命令替换"},
    {"id": "CMD-003", "payload": "| cat /etc/passwd", "description": "命令注入-管道"},
]

ALL_SECURITY_PAYLOADS = SQL_INJECTION_PAYLOADS + XSS_PAYLOADS + PATH_TRAVERSAL_PAYLOADS + COMMAND_INJECTION_PAYLOADS

# 依赖参数识别模式
DEPENDENCY_PATTERNS = {
    "TOKEN": {
        "matchers": [
            {"param_location": "header", "param_name_pattern": r"(?i)^authorization$"},
            {"param_location": "header", "param_name_pattern": r"(?i)^x-auth"},
            {"param_location": "header", "param_name_pattern": r"(?i)^token$"},
        ],
        "source_api_pattern": r"POST_.*/(auth|login|token)",
        "extract_path": "data.token",
        "description": "登录获取 Token",
    },
}

# 资源 ID 依赖识别
ID_PARAM_PATTERNS = [
    (r"(?i)^(id|userId|uid)$", "USER_ID", r"POST_.*/(user|auth/register)"),
    (r"(?i)^(orderId|order_id)$", "ORDER_ID", r"POST_.*/order$"),
    (r"(?i)^(addressId|address_id)$", "ADDRESS_ID", r"POST_.*/address$"),
    (r"(?i)^(productId|product_id)$", "PRODUCT_ID", r"POST_.*/(product|admin/product)"),
    (r"(?i)^(cartId|cart_id)$", "CART_ID", r"POST_.*/cart$"),
    (r"(?i)^(categoryId|category_id)$", "CATEGORY_ID", r"POST_.*/(category|admin/category)"),
    (r"(?i)^(bannerId|banner_id)$", "BANNER_ID", r"POST_.*/(banner|admin/banner)"),
    (r"(?i)^(captchaKey|captcha_key|key)$", "CAPTCHA_KEY", r"GET_.*/captcha"),
    (r"(?i)^(cartIds|cart_ids)$", "CART_IDS", r"POST_.*/cart"),
]

# ============================================================================
# 自然语言模式 - 字段类型定义
# ============================================================================

# 关键词 → 字段标识 映射（按优先级排列，更具体的关键词在前）
NL_FIELD_KEYWORDS = [
    # 身份证
    (["身份证号", "身份证号码", "身份证", "证件号", "ID号"], "ssn"),
    # 手机号
    (["手机号码", "手机号", "联系电话", "电话号码", "手机", "电话", "电话号"], "phone"),
    # 用户名
    (["用户名", "账号", "账户名", "登录名"], "username"),
    # 姓名
    (["中文姓名", "真实姓名", "姓名", "名字"], "name"),
    # 邮箱
    (["电子邮件", "邮箱", "email", "E-mail", "邮件"], "email"),
    # 地址
    (["收货地址", "通讯地址", "详细地址", "地址", "住址"], "address"),
    # 公司名
    (["公司名称", "公司名", "企业名", "企业", "公司"], "company"),
    # 密码
    (["密码", "口令"], "password"),
    # 日期时间
    (["日期时间", "时间戳", "时间"], "datetime"),
    # 日期
    (["日期"], "date"),
    # URL
    (["网址", "链接", "网站", "URL", "url"], "url"),
    # IP
    (["IP地址", "IP", "ipv4", "IP地址"], "ipv4"),
    # 银行卡
    (["银行卡号", "银行卡", "信用卡号", "信用卡"], "credit_card"),
    # 城市
    (["城市名", "城市"], "city"),
    # 省份
    (["省份", "省"], "province"),
    # 邮编
    (["邮政编码", "邮编"], "postcode"),
    # 职业
    (["职位", "职务", "职业"], "job"),
    # 车牌号
    (["车牌号", "车牌"], "license_plate"),
    # 性别
    (["性别"], "gender"),
    # 年龄
    (["年龄"], "age"),
]

# 字段标识 → 中文显示名
NL_FIELD_DISPLAY_NAMES = {
    "ssn": "身份证号",
    "phone": "手机号",
    "username": "用户名",
    "name": "姓名",
    "email": "邮箱",
    "address": "地址",
    "company": "公司名",
    "password": "密码",
    "date": "日期",
    "datetime": "日期时间",
    "url": "URL",
    "ipv4": "IP地址",
    "credit_card": "银行卡号",
    "city": "城市",
    "province": "省份",
    "postcode": "邮编",
    "job": "职业",
    "license_plate": "车牌号",
    "gender": "性别",
    "age": "年龄",
}

# 默认字段集合（当用户描述模糊时使用）
NL_DEFAULT_FIELD_SETS = {
    "用户数据": ["name", "phone", "email", "username", "gender", "age"],
    "用户完整信息": ["username", "name", "phone", "email", "ssn", "address", "password", "gender", "age"],
    "注册数据": ["username", "password", "email", "phone"],
    "收货数据": ["name", "phone", "address", "province", "city", "postcode"],
    "地址数据": ["name", "phone", "address", "province", "city", "postcode"],
    "公司数据": ["company", "address", "city", "phone", "email"],
    "企业数据": ["company", "address", "city", "phone", "email"],
    "测试数据": ["name", "phone", "email", "address"],
    "性能测试数据": ["username", "password", "email", "phone"],
}

# 中国手机号合法号段前3位
CHINA_MOBILE_PREFIXES = [
    "134", "135", "136", "137", "138", "139", "147", "150", "151", "152",
    "157", "158", "159", "172", "178", "182", "183", "184", "187", "188",
    "195", "197", "198",
]
CHINA_UNICOM_PREFIXES = [
    "130", "131", "132", "145", "155", "156", "166", "171", "175", "176",
    "185", "186", "196",
]
CHINA_TELECOM_PREFIXES = [
    "133", "149", "153", "173", "174", "177", "180", "181", "189",
    "190", "191", "193", "199",
]
ALL_PHONE_PREFIXES = CHINA_MOBILE_PREFIXES + CHINA_UNICOM_PREFIXES + CHINA_TELECOM_PREFIXES

# 身份证地区码（常用）
ID_AREA_CODES = [
    "110101", "110102", "110105", "110106",  # 北京
    "310115", "310101", "310104", "310105",  # 上海
    "440305", "440304", "440306", "440303",  # 深圳
    "440106", "440103", "440104", "440105",  # 广州
    "330102", "330106", "330108", "330109",  # 杭州
    "510107", "510104", "510105", "510106",  # 成都
    "320102", "320104", "320105", "320106",  # 南京
    "420102", "420103", "420104", "420106",  # 武汉
]

# 百家姓（常用）
COMMON_SURNAMES = list("赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜戚谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳酆鲍史唐费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮齐康伍余元卜顾孟平黄和穆萧尹姚邵湛汪祁毛禹狄米贝明臧计伏成戴谈宋茅庞熊纪舒屈项祝董梁杜阮蓝闵席季麻强贾路娄危江童颜郭梅盛林刁钟徐邱骆高夏蔡田樊胡凌霍虞万支柯昝管卢莫经房裘缪干解应宗丁宣贲邓郁单杭洪包诸左石崔吉钮龚程嵇邢滑裴陆荣翁荀羊於惠甄曲家封芮羿储靳汲邴糜松井段富巫乌焦巴弓牧隗山谷车侯宓蓬全郗班仰秋仲伊宫宁仇栾暴甘钭厉戎祖武符刘景詹束龙叶幸司韶郜黎蓟薄印宿白怀蒲邰从鄂索咸籍赖卓蔺屠蒙池乔阴郁胥能苍双闻莘党翟谭贡劳逄姬申扶堵冉宰郦雍却璩桑桂濮牛寿通边扈燕冀郏浦尚农温别庄晏柴瞿阎充慕连茹习宦艾鱼容向古易慎戈廖庾终暨居衡步都耿满弘匡国文寇广禄阙东欧殳沃利蔚越夔隆师巩厍聂晁勾敖融冷訾辛阚那简饶空曾毋沙乜养鞠须丰巢关蒯相查后荆红游竺权逯盖益桓公")

# 常用名字
COMMON_GIVEN_NAMES = [
    "伟", "芳", "娜", "秀英", "敏", "静", "丽", "强", "磊", "军",
    "洋", "勇", "艳", "杰", "娟", "涛", "明", "超", "秀兰", "霞",
    "平", "刚", "桂英", "文", "华", "慧", "建", "国", "小红", "志",
    "金", "玲", "彬", "婷", "莉", "鹏", "宇", "浩", "子涵", "欣怡",
    "思远", "梦琪", "佳怡", "浩然", "子轩", "梓涵", "一诺", "语桐", "泽宇", "诗涵",
]


# ============================================================================
# Faker 封装（可选依赖，不可用时回退到内置生成器）
# ============================================================================

_faker_zh = None
_faker_en = None
_faker_available = False

def _init_faker():
    """初始化 Faker 实例（延迟加载）"""
    global _faker_zh, _faker_en, _faker_available
    if _faker_available:
        return
    try:
        from faker import Faker
        _faker_zh = Faker('zh_CN')
        _faker_en = Faker('en_US')
        _faker_available = True
    except ImportError:
        _faker_available = False


def _generate_faker_value(field_type: str) -> str:
    """使用 Faker 生成合法数据"""
    _init_faker()

    if _faker_available:
        try:
            if field_type == "ssn":
                return _faker_zh.ssn()
            elif field_type == "phone":
                return _faker_zh.phone_number()
            elif field_type == "username":
                return _faker_en.user_name()
            elif field_type == "name":
                return _faker_zh.name()
            elif field_type == "email":
                return _faker_en.email()
            elif field_type == "address":
                return _faker_zh.address().replace("\n", "")
            elif field_type == "company":
                return _faker_zh.company()
            elif field_type == "date":
                return str(_faker_zh.date())
            elif field_type == "datetime":
                return str(_faker_zh.date_time())
            elif field_type == "url":
                return _faker_en.url()
            elif field_type == "ipv4":
                return _faker_en.ipv4()
            elif field_type == "credit_card":
                return _faker_en.credit_card_number()
            elif field_type == "city":
                return _faker_zh.city()
            elif field_type == "province":
                return _faker_zh.province()
            elif field_type == "postcode":
                return _faker_zh.postcode()
            elif field_type == "job":
                return _faker_zh.job()
            elif field_type == "password":
                return _generate_password_builtin()
            elif field_type == "license_plate":
                return _generate_license_plate_builtin()
            elif field_type == "gender":
                return random.choice(["男", "女"])
            elif field_type == "age":
                return str(random.randint(18, 65))
            else:
                return _faker_en.text(max_nb_chars=20).strip()
        except Exception:
            pass

    # Faker 不可用，回退到内置生成器
    return _generate_builtin_value(field_type)


def _generate_password_builtin() -> str:
    """内置密码生成器"""
    length = random.randint(8, 16)
    chars = [
        random.choice(string.ascii_uppercase),
        random.choice(string.ascii_lowercase),
        random.choice(string.digits),
        random.choice("!@#$%^&*"),
    ]
    remaining = ''.join(random.choices(
        string.ascii_letters + string.digits + "!@#$%^&*", k=max(0, length - 4)
    ))
    password = list(chars[0] + chars[1] + chars[2] + chars[3] + remaining)
    random.shuffle(password)
    return ''.join(password)


def _generate_license_plate_builtin() -> str:
    """内置车牌号生成器"""
    provinces = "京沪津渝冀豫云辽黑湘皖鲁新苏浙赣鄂桂甘晋蒙陕吉闽贵粤川青藏琼宁"
    letter = random.choice(string.ascii_uppercase)
    suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
    return f"{random.choice(provinces)}{letter}·{suffix}"


def _compute_id_checksum(id17: str) -> str:
    """计算中国身份证校验位"""
    weights = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
    check_codes = ['1', '0', 'X', '9', '8', '7', '6', '5', '4', '3', '2']
    total = sum(int(id17[i]) * weights[i] for i in range(17))
    return check_codes[total % 11]


def _generate_ssn_builtin() -> str:
    """内置身份证号生成器"""
    area_code = random.choice(ID_AREA_CODES)
    year = random.randint(1960, 2005)
    month = random.randint(1, 12)
    max_day = 31 if month in [1,3,5,7,8,10,12] else (30 if month in [4,6,9,11] else (29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28))
    day = random.randint(1, max_day)
    seq = random.randint(10, 99)
    id17 = f"{area_code}{year:04d}{month:02d}{day:02d}{seq:03d}"[:17]
    checksum = _compute_id_checksum(id17)
    return id17 + checksum


def _generate_phone_builtin() -> str:
    """内置手机号生成器"""
    prefix = random.choice(ALL_PHONE_PREFIXES)
    suffix = ''.join(random.choices(string.digits, k=8))
    return prefix + suffix


def _generate_name_builtin() -> str:
    """内置中文姓名生成器"""
    surname = random.choice(COMMON_SURNAMES)
    given_name = random.choice(COMMON_GIVEN_NAMES)
    return surname + given_name


def _generate_address_builtin() -> str:
    """内置中文地址生成器"""
    provinces_cities = [
        ("北京市", "朝阳区"), ("北京市", "海淀区"), ("北京市", "东城区"),
        ("上海市", "浦东新区"), ("上海市", "黄浦区"), ("上海市", "徐汇区"),
        ("广东省深圳市", "南山区"), ("广东省深圳市", "福田区"),
        ("广东省广州市", "天河区"), ("浙江省杭州市", "上城区"),
        ("四川省成都市", "武侯区"), ("江苏省南京市", "鼓楼区"),
    ]
    province_city = random.choice(provinces_cities)
    streets = ["建国路", "长安街", "中关村大街", "陆家嘴环路", "科技园路",
               "人民路", "中山路", "解放路", "和平路", "幸福路"]
    street = random.choice(streets)
    number = random.randint(1, 999)
    return f"{province_city[0]}{province_city[1]}{street}{number}号"


def _generate_company_builtin() -> str:
    """内置公司名生成器"""
    cities = ["北京", "上海", "深圳", "广州", "杭州", "成都", "南京", "武汉"]
    names = ["创新", "瑞达", "远景", "华信", "中科", "博远", "天宇", "启明", "恒通", "智汇"]
    industries = ["科技", "贸易", "网络", "信息", "教育", "咨询", "传媒", "数据"]
    suffixes = ["有限公司", "股份有限公司", "科技有限公司", "集团有限公司"]
    return f"{random.choice(cities)}{random.choice(names)}{random.choice(industries)}{random.choice(suffixes)}"


def _generate_builtin_value(field_type: str) -> str:
    """内置数据生成器（Faker 不可用时的回退方案）"""
    generators = {
        "ssn": _generate_ssn_builtin,
        "phone": _generate_phone_builtin,
        "username": lambda: f"user_{_generate_random_string(6, string.ascii_lowercase + string.digits)}",
        "name": _generate_name_builtin,
        "email": lambda: f"test{_generate_random_string(4, string.digits)}@example.com",
        "address": _generate_address_builtin,
        "company": _generate_company_builtin,
        "password": _generate_password_builtin,
        "date": lambda: f"{random.randint(2020,2026)}-{random.randint(1,12):02d}-{random.randint(1,28):02d}",
        "datetime": lambda: f"{random.randint(2020,2026)}-{random.randint(1,12):02d}-{random.randint(1,28):02d}T{random.randint(0,23):02d}:{random.randint(0,59):02d}:{random.randint(0,59):02d}+08:00",
        "url": lambda: f"https://example.com/{_generate_random_string(6, string.ascii_lowercase)}",
        "ipv4": lambda: f"{random.randint(1,255)}.{random.randint(0,255)}.{random.randint(0,255)}.{random.randint(1,255)}",
        "credit_card": lambda: ''.join(random.choices(string.digits, k=16)),
        "city": lambda: random.choice(["北京", "上海", "深圳", "广州", "杭州", "成都", "南京", "武汉", "西安", "重庆"]),
        "province": lambda: random.choice(["北京市", "上海市", "广东省", "浙江省", "江苏省", "四川省", "湖北省", "陕西省", "重庆市"]),
        "postcode": lambda: ''.join(random.choices(string.digits, k=6)),
        "job": lambda: random.choice(["软件工程师", "产品经理", "测试工程师", "设计师", "运营", "市场经理", "人事专员", "财务", "法务", "行政"]),
        "license_plate": _generate_license_plate_builtin,
        "gender": lambda: random.choice(["男", "女"]),
        "age": lambda: str(random.randint(18, 65)),
    }
    gen = generators.get(field_type)
    if gen:
        return gen()
    return _generate_random_string(10)


# ============================================================================
# 非法数据生成器
# ============================================================================

def _generate_invalid_ssn() -> List[Tuple[str, str]]:
    """生成非法身份证号变体"""
    valid = _generate_faker_value("ssn")
    return [
        (valid[:17], "17位-少校验位"),
        (valid + "1", "19位-多1位"),
        (valid[:10] + "A" + valid[11:18], "含字母"),
        (valid[:17] + _get_wrong_checksum(valid[:17]), "校验位错误"),
        ("", "空字符串"),
        ("110101199002301234", "非法日期-2月30日"),
    ]


def _get_wrong_checksum(id17: str) -> str:
    """获取错误的校验位"""
    correct = _compute_id_checksum(id17)
    wrong_chars = [c for c in "0123456789X" if c != correct]
    return random.choice(wrong_chars)


def _generate_invalid_phone() -> List[Tuple[str, str]]:
    """生成非法手机号变体"""
    return [
        ("1380000123", "10位-少1位"),
        ("138000012345", "12位-多1位"),
        ("23800001234", "非1开头"),
        ("1380000abcd", "含字母"),
        ("138-000-1234", "含特殊字符"),
        ("", "空字符串"),
    ]


def _generate_invalid_email() -> List[Tuple[str, str]]:
    """生成非法邮箱变体"""
    return [
        ("testexample.com", "缺@符号"),
        ("test@", "缺域名"),
        ("@example.com", "缺用户名"),
        ("test..name@example.com", "连续点号"),
        ("test @example.com", "含空格"),
        ("test@example", "无顶级域名"),
    ]


def _generate_invalid_url() -> List[Tuple[str, str]]:
    """生成非法URL变体"""
    return [
        ("example.com", "缺协议"),
        ("ftp://example.com", "非HTTP协议"),
        ("https://example .com", "含空格"),
        ("://example.com", "缺协议名"),
    ]


def _generate_invalid_ipv4() -> List[Tuple[str, str]]:
    """生成非法IP变体"""
    return [
        ("192.168.1", "段数不够"),
        ("256.1.1.1", "段值超范围"),
        ("192.168.1.abc", "含字母"),
        ("192.168.-1.1", "含负数"),
    ]


def _generate_invalid_password() -> List[Tuple[str, str]]:
    """生成非法密码变体"""
    return [
        ("Test1", "少于8位"),
        ("12345678", "纯数字"),
        ("abcdefgh", "纯字母"),
        ("Test1234", "无特殊字符"),
        ("", "空字符串"),
    ]


def _generate_invalid_username() -> List[Tuple[str, str]]:
    """生成非法用户名变体"""
    return [
        ("test@user", "含特殊字符"),
        ("test user", "含空格"),
        ("a", "过短-1位"),
        ("a" * 25, "过长-25位"),
    ]


def _generate_invalid_value(field_type: str) -> List[Tuple[str, str]]:
    """为指定字段类型生成非法数据变体列表"""
    invalid_generators = {
        "ssn": _generate_invalid_ssn,
        "phone": _generate_invalid_phone,
        "email": _generate_invalid_email,
        "url": _generate_invalid_url,
        "ipv4": _generate_invalid_ipv4,
        "password": _generate_invalid_password,
        "username": _generate_invalid_username,
    }

    # 数值型字段的非法变体
    if field_type == "age":
        return [
            ("-1", "负数"),
            ("0", "零值"),
            ("999", "超大数值"),
            ("17.5", "小数"),
            ("abc", "非数字"),
            ("", "空字符串"),
        ]
    if field_type == "gender":
        return [
            ("", "空字符串"),
            ("3", "非枚举数值"),
            ("unknown", "非枚举字符串"),
            ("null", "null值"),
        ]

    generator = invalid_generators.get(field_type)
    if generator:
        return generator()

    # 通用非法变体
    return [
        ("", "空字符串"),
        ("   ", "纯空格"),
        ("null", "null值"),
        ("' OR '1'='1", "SQL注入"),
        ("<script>alert(1)</script>", "XSS注入"),
    ]


# ============================================================================
# 自然语言解析器
# ============================================================================

class NLParser:
    """自然语言描述解析器，提取字段类型、数量、合法性约束"""

    @staticmethod
    def parse(prompt: str) -> List[dict]:
        """解析自然语言描述，返回字段需求列表

        返回格式:
        [
            {
                "field_type": "phone",     # 字段标识
                "display_name": "手机号",    # 显示名
                "count": 5,                 # 数量
                "validity": "valid",        # valid/invalid/boundary
            },
            ...
        ]
        """
        results = []
        prompt_lower = prompt.lower()

        # 0. 先从整体 prompt 提取全局数量，用于子句数量继承
        global_count = NLParser._extract_count(prompt)
        global_validity = NLParser._extract_validity(prompt)

        # 1. 尝试按"和""与"分割子句
        clauses = re.split(r'[，,、和与以及还有；;]', prompt)

        # 收集"包含"子句中的字段
        contain_fields = []
        # 跟踪默认字段集匹配产生的结果索引（用于"包含"替换）
        default_set_result_indices = []

        for clause in clauses:
            clause = clause.strip()
            if not clause:
                continue

            # 提取数量（子句没有数量时继承全局数量）
            clause_count = NLParser._extract_count(clause)
            count = clause_count if NLParser._has_explicit_count(clause) else global_count

            # 提取合法性
            validity = NLParser._extract_validity(clause)

            # "包含"子句：收集字段但不立即添加到results
            if clause.startswith("包含"):
                field_types = NLParser._extract_field_types(clause)
                if field_types:
                    contain_fields.extend(field_types)
                continue

            # 提取字段类型
            field_types = NLParser._extract_field_types(clause)

            if field_types:
                for ft in field_types:
                    results.append({
                        "field_type": ft,
                        "display_name": NL_FIELD_DISPLAY_NAMES.get(ft, ft),
                        "count": count,
                        "validity": validity,
                    })
            else:
                # 尝试默认字段集
                for key, fields in NL_DEFAULT_FIELD_SETS.items():
                    if key in clause:
                        start_idx = len(results)
                        for ft in fields:
                            results.append({
                                "field_type": ft,
                                "display_name": NL_FIELD_DISPLAY_NAMES.get(ft, ft),
                                "count": count,
                                "validity": validity,
                            })
                            default_set_result_indices.append(start_idx)
                            start_idx += 1
                        break

        # 2. 如果有"包含"子句指定的字段，替换默认字段集的结果
        if contain_fields and default_set_result_indices:
            # 移除默认字段集产生的结果（按索引倒序删除）
            for idx in sorted(default_set_result_indices, reverse=True):
                if idx < len(results):
                    results.pop(idx)
            # 添加"包含"子句指定的字段
            for ft in contain_fields:
                results.append({
                    "field_type": ft,
                    "display_name": NL_FIELD_DISPLAY_NAMES.get(ft, ft),
                    "count": global_count,
                    "validity": global_validity,
                })

        # 3. 如果没有解析结果，尝试整体解析
        if not results:
            field_types = NLParser._extract_field_types(prompt)

            if field_types:
                for ft in field_types:
                    results.append({
                        "field_type": ft,
                        "display_name": NL_FIELD_DISPLAY_NAMES.get(ft, ft),
                        "count": global_count,
                        "validity": global_validity,
                    })
            else:
                # 默认字段集
                for key, fields in NL_DEFAULT_FIELD_SETS.items():
                    if key in prompt:
                        for ft in fields:
                            results.append({
                                "field_type": ft,
                                "display_name": NL_FIELD_DISPLAY_NAMES.get(ft, ft),
                                "count": global_count,
                                "validity": global_validity,
                            })
                        break
                else:
                    # 最终回退：生成基础字段
                    for ft in ["name", "phone", "email"]:
                        results.append({
                            "field_type": ft,
                            "display_name": NL_FIELD_DISPLAY_NAMES.get(ft, ft),
                            "count": global_count,
                            "validity": global_validity,
                        })

        return results

    @staticmethod
    def _has_explicit_count(text: str) -> bool:
        """判断文本中是否显式包含数量词"""
        cn_nums = {"一", "二", "两", "三", "四", "五", "六", "七", "八", "九", "十", "百", "千", "万"}
        for cn in cn_nums:
            if re.search(f'{cn}(?:个|条|组|份|行|笔)', text):
                return True
        if re.search(r'\d+\s*(?:个|条|组|份|行|笔|条数据|组数据)', text):
            return True
        # 排除纯数字但非数量词的情况（如"年龄"中的"龄"不含数字）
        # 如果文本中有"数字+量词"模式，则认为是显式数量
        return False

    @staticmethod
    def _extract_count(text: str) -> int:
        """从文本中提取数量"""
        # 匹配中文数字
        cn_nums = {"一": 1, "二": 2, "两": 2, "三": 3, "四": 4, "五": 5,
                   "六": 6, "七": 7, "八": 8, "九": 9, "十": 10,
                   "百": 100, "千": 1000, "万": 10000}
        for cn, num in cn_nums.items():
            m = re.search(f'{cn}(?:个|条|组|份|行|笔)', text)
            if m:
                return num

        # 匹配阿拉伯数字
        m = re.search(r'(\d+)\s*(?:个|条|组|份|行|笔|条数据|组数据)', text)
        if m:
            return int(m.group(1))

        # 匹配纯数字
        m = re.search(r'(\d+)', text)
        if m:
            return int(m.group(1))

        # 默认
        if "若干" in text or "一些" in text or "几个" in text:
            return 5

        return 5  # 默认5条

    @staticmethod
    def _extract_validity(text: str) -> str:
        """从文本中提取合法性要求"""
        if re.search(r'非法|无效|错误|异常|不合法|不正确|invalid', text, re.IGNORECASE):
            return "invalid"
        if re.search(r'边界|临界|极限|boundary', text, re.IGNORECASE):
            return "boundary"
        if re.search(r'合法|有效|正确|正常|valid|符合', text, re.IGNORECASE):
            return "valid"
        return "valid"  # 默认合法

    @staticmethod
    def _extract_field_types(text: str) -> List[str]:
        """从文本中提取字段类型"""
        found = []
        matched_spans = []  # 记录已匹配的文本区间，避免重复匹配

        for keywords, field_type in NL_FIELD_KEYWORDS:
            for kw in keywords:
                # 找到关键词的位置
                for m in re.finditer(re.escape(kw), text):
                    start, end = m.start(), m.end()
                    # 检查是否与已匹配的区间重叠
                    overlaps = False
                    for ms, me in matched_spans:
                        if not (end <= ms or start >= me):
                            # 有重叠，只有当当前关键词更长时替换
                            if end - start > me - ms:
                                matched_spans.remove((ms, me))
                                found = [ft for ft in found if ft != field_type]
                            else:
                                overlaps = True
                                break
                    if not overlaps:
                        matched_spans.append((start, end))
                        if field_type not in found:
                            found.append(field_type)
                    break  # 每个关键词只匹配一次
                else:
                    continue
                break  # 如果已匹配，跳到下一个字段类型

        return found


# ============================================================================
# 自然语言模式 - 数据生成
# ============================================================================

def generate_nl_data(prompt: str, output_dir: str) -> dict:
    """自然语言模式主生成函数

    Args:
        prompt: 自然语言描述
        output_dir: 输出目录

    Returns:
        汇总信息字典
    """
    # 解析描述
    field_requests = NLParser.parse(prompt)

    if not field_requests:
        print(f"无法解析描述：{prompt}", file=sys.stderr)
        print("支持的字段类型：身份证号、手机号、用户名、姓名、邮箱、地址、公司名、"
              "密码、日期、日期时间、URL、IP地址、银行卡号、城市、省份、邮编、职业、车牌号、性别、年龄",
              file=sys.stderr)
        return {"error": "无法解析描述"}

    # 生成数据
    all_data = {}  # {field_type: [values]}
    for req in field_requests:
        ft = req["field_type"]
        count = req["count"]
        validity = req["validity"]

        if ft not in all_data:
            all_data[ft] = {"valid": [], "invalid": [], "boundary": []}

        if validity == "valid":
            for _ in range(count):
                all_data[ft]["valid"].append(_generate_faker_value(ft))
        elif validity == "invalid":
            invalid_variants = _generate_invalid_value(ft)
            # 按需求数量从变体中选择
            for i in range(min(count, len(invalid_variants))):
                all_data[ft]["invalid"].append(invalid_variants[i])
            # 如果需要更多非法数据，循环使用变体
            while len(all_data[ft]["invalid"]) < count:
                idx = len(all_data[ft]["invalid"]) % len(invalid_variants)
                all_data[ft]["invalid"].append(invalid_variants[idx])
        elif validity == "boundary":
            # 边界数据：使用合法数据的极端形式
            for _ in range(count):
                all_data[ft]["boundary"].append(_generate_faker_value(ft))

    # 输出 CSV
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now(timezone(timedelta(hours=8))).strftime("%Y%m%d_%H%M%S")

    # 收集所有字段
    field_types = list(all_data.keys())
    display_names = [NL_FIELD_DISPLAY_NAMES.get(ft, ft) for ft in field_types]

    # 判断是否需要区分合法/非法
    has_invalid = any(len(all_data[ft]["invalid"]) > 0 for ft in field_types)
    has_boundary = any(len(all_data[ft]["boundary"]) > 0 for ft in field_types)

    if has_invalid or has_boundary:
        # 需要区分合法性，添加标记列
        _write_nl_csv_with_validity(all_data, field_types, display_names, output_dir, timestamp)
    else:
        # 纯合法数据，简洁输出
        _write_nl_csv_simple(all_data, field_types, display_names, output_dir, timestamp)

    # 也输出 JSON 汇总
    summary = {
        "mode": "nl",
        "prompt": prompt,
        "fields": field_types,
        "total_rows": sum(
            len(all_data[ft].get("valid", [])) +
            len(all_data[ft].get("invalid", [])) +
            len(all_data[ft].get("boundary", []))
            for ft in field_types
        ),
        "stats": {
            ft: {
                "valid": len(all_data[ft].get("valid", [])),
                "invalid": len(all_data[ft].get("invalid", [])),
                "boundary": len(all_data[ft].get("boundary", [])),
            }
            for ft in field_types
        },
    }
    _write_json(summary, os.path.join(output_dir, f"nl_summary_{timestamp}.json"))

    return summary


def _write_nl_csv_simple(all_data: dict, field_types: list, display_names: list,
                          output_dir: str, timestamp: str):
    """写入简单格式 CSV（纯合法数据）"""
    filename = f"nl_test_data_{timestamp}.csv"
    filepath = os.path.join(output_dir, filename)

    # 计算最大行数
    max_rows = max(len(all_data[ft].get("valid", [])) for ft in field_types)

    with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(display_names)
        for i in range(max_rows):
            row = []
            for ft in field_types:
                valid_list = all_data[ft].get("valid", [])
                row.append(valid_list[i] if i < len(valid_list) else "")
            writer.writerow(row)

    print(f"  输出文件: {filepath}")


def _write_nl_csv_with_validity(all_data: dict, field_types: list, display_names: list,
                                 output_dir: str, timestamp: str):
    """写入带合法性标记的 CSV"""
    filename = f"nl_test_data_{timestamp}.csv"
    filepath = os.path.join(output_dir, filename)

    with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)

        # 按字段分组写入
        for idx, ft in enumerate(field_types):
            dn = display_names[idx]

            # 写入字段标题行
            writer.writerow([f"=== {dn} ==="])

            # 合法数据
            valid_list = all_data[ft].get("valid", [])
            if valid_list:
                writer.writerow([dn, "合法性", "说明"])
                for val in valid_list:
                    writer.writerow([val, "合法", ""])
                writer.writerow([])  # 空行分隔

            # 非法数据
            invalid_list = all_data[ft].get("invalid", [])
            if invalid_list:
                writer.writerow([dn, "合法性", "说明"])
                for val, desc in invalid_list:
                    writer.writerow([val, "非法", desc])
                writer.writerow([])  # 空行分隔

            # 边界数据
            boundary_list = all_data[ft].get("boundary", [])
            if boundary_list:
                writer.writerow([dn, "合法性", "说明"])
                for val in boundary_list:
                    writer.writerow([val, "边界", ""])
                writer.writerow([])  # 空行分隔

    print(f"  输出文件: {filepath}")


# ============================================================================
# 规范模式 - 数据生成
# ============================================================================

def generate_spec_data(spec_path: str, output_dir: str, output_format: str = "csv",
                       dimensions: list = None) -> dict:
    """规范模式主生成函数

    Args:
        spec_path: 字段规则文件路径
        output_dir: 输出目录
        output_format: 输出格式
        dimensions: 生成维度

    Returns:
        汇总信息字典
    """
    # 加载规则
    with open(spec_path, 'r', encoding='utf-8') as f:
        if spec_path.endswith('.json'):
            spec = json.load(f)
        else:
            try:
                import yaml
                spec = yaml.safe_load(f)
            except ImportError:
                print("错误: YAML 格式需要安装 PyYAML 库", file=sys.stderr)
                sys.exit(1)

    fields = spec.get("fields", [])
    spec_name = spec.get("name", "spec_data")
    default_count = spec.get("count", 10)
    spec_dimensions = spec.get("dimensions", ["positive", "boundary", "negative"])

    if not dimensions:
        dimensions = spec_dimensions

    if not fields:
        print("错误: 规则文件中未定义 fields", file=sys.stderr)
        sys.exit(1)

    os.makedirs(output_dir, exist_ok=True)

    # 为每个维度生成数据
    all_results = {}
    stats = {"total_rows": 0, "by_dimension": {}}

    for dim in dimensions:
        if dim == "positive":
            rows = _generate_spec_positive(fields, default_count)
        elif dim == "boundary":
            rows = _generate_spec_boundary(fields, default_count)
        elif dim == "negative":
            rows = _generate_spec_negative(fields, default_count)
        else:
            continue

        all_results[dim] = rows
        stats["by_dimension"][dim] = len(rows)
        stats["total_rows"] += len(rows)

    # 输出
    if output_format == "csv":
        _write_spec_csv(all_results, fields, output_dir, spec_name)
    elif output_format == "json":
        _write_spec_json(all_results, fields, output_dir, spec_name)
    elif output_format == "yaml":
        _write_spec_yaml(all_results, fields, output_dir, spec_name)
    elif output_format == "excel":
        _write_spec_excel(all_results, fields, output_dir, spec_name)

    # 汇总
    summary = {
        "mode": "spec",
        "spec_name": spec_name,
        "field_count": len(fields),
        "stats": stats,
    }
    _write_json(summary, os.path.join(output_dir, f"spec_summary_{spec_name}.json"))

    return summary


def _generate_spec_positive(fields: list, count: int) -> List[dict]:
    """生成规范模式的正向数据"""
    rows = []
    for i in range(count):
        row = {}
        for field in fields:
            name = field.get("name", "")
            f_type = field.get("type", "string")
            row[name] = _generate_spec_field_value(field, "positive")
        rows.append(row)
    return rows


def _generate_spec_boundary(fields: list, count: int) -> List[dict]:
    """生成规范模式的边界值数据"""
    rows = []
    for field in fields:
        name = field.get("name", "")
        f_type = field.get("type", "string")

        boundary_values = _generate_spec_boundary_for_field(field)
        for val, desc in boundary_values:
            row = {}
            for f in fields:
                fn = f.get("name", "")
                if fn == name:
                    row[fn] = val
                else:
                    row[fn] = _generate_spec_field_value(f, "positive")
            row["_boundary_field"] = name
            row["_boundary_desc"] = desc
            rows.append(row)

    return rows


def _generate_spec_negative(fields: list, count: int) -> List[dict]:
    """生成规范模式的异常数据"""
    rows = []
    for field in fields:
        name = field.get("name", "")
        required = field.get("required", False)
        f_type = field.get("type", "string")

        negative_values = _generate_spec_negative_for_field(field)
        for val, desc in negative_values:
            row = {}
            for f in fields:
                fn = f.get("name", "")
                if fn == name:
                    row[fn] = val
                else:
                    row[fn] = _generate_spec_field_value(f, "positive")
            row["_negative_field"] = name
            row["_negative_desc"] = desc
            rows.append(row)

    return rows


def _generate_spec_field_value(field: dict, category: str = "positive") -> Any:
    """为规范模式字段生成单个值"""
    name = field.get("name", "").lower()
    f_type = field.get("type", "string")
    f_format = field.get("format", "")
    f_pattern = field.get("pattern", "")
    f_enum = field.get("enum", None)
    f_default = field.get("default", None)
    f_min_len = field.get("minLength", None)
    f_max_len = field.get("maxLength", None)
    f_min = field.get("minimum", None)
    f_max = field.get("maximum", None)

    # 有默认值
    if f_default is not None:
        return f_default

    # 有枚举
    if f_enum:
        return random.choice(f_enum)

    # 按 format 生成
    format_map = {
        "email": "email",
        "uri": "url", "url": "url",
        "phone": "phone",
        "date": "date",
        "date-time": "datetime",
        "uuid": None,
    }
    if f_format in format_map and format_map[f_format]:
        return _generate_faker_value(format_map[f_format])

    # 按名称推断
    name_field_map = {
        "username": "username", "email": "email", "phone": "phone",
        "mobile": "phone", "password": "password", "address": "address",
        "url": "url", "name": "name",
    }
    for key, ft in name_field_map.items():
        if key in name:
            return _generate_faker_value(ft)

    # 按类型生成
    if f_type == "string":
        min_len = f_min_len or 1
        max_len = f_max_len or 255
        length = min((min_len + max_len) // 2, 20)
        if f_pattern:
            return _generate_valid_value_for_pattern(f_pattern, length)
        return _generate_random_string(max(length, 5))
    elif f_type == "integer":
        minimum = f_min if f_min is not None else 1
        maximum = f_max if f_max is not None else 1000
        return (minimum + maximum) // 2
    elif f_type == "number":
        minimum = f_min if f_min is not None else 0
        maximum = f_max if f_max is not None else 1000
        return round((minimum + maximum) / 2, 2)
    elif f_type == "boolean":
        return random.choice([True, False])
    elif f_type == "array":
        return []
    elif f_type == "object":
        return {}
    return "test_value"


def _generate_spec_boundary_for_field(field: dict) -> List[Tuple[Any, str]]:
    """为规范模式字段生成边界值"""
    results = []
    f_type = field.get("type", "string")
    f_min_len = field.get("minLength")
    f_max_len = field.get("maxLength")
    f_min = field.get("minimum")
    f_max = field.get("maximum")
    f_pattern = field.get("pattern", "")
    f_enum = field.get("enum")

    # 从正则中提取隐含长度约束
    if f_pattern:
        inferred_min, inferred_max = _extract_length_from_pattern(f_pattern)
        if inferred_min is not None and (f_min_len is None or inferred_min > f_min_len):
            f_min_len = inferred_min
        if inferred_max is not None and (f_max_len is None or inferred_max < f_max_len):
            f_max_len = inferred_max

    if f_type == "string":
        if f_min_len is not None and f_min_len > 0:
            results.append((_generate_string_by_length(f_min_len - 1, f_pattern or None),
                          f"最小长度-1={f_min_len - 1}(非法)"))
            results.append((_generate_string_by_length(f_min_len, f_pattern or None),
                          f"最小长度={f_min_len}(合法)"))
        if f_max_len is not None and f_max_len < 10000:
            results.append((_generate_string_by_length(f_max_len, f_pattern or None),
                          f"最大长度={f_max_len}(合法)"))
            results.append((_generate_string_by_length(f_max_len + 1, f_pattern or None),
                          f"最大长度+1={f_max_len + 1}(非法)"))
    elif f_type in ("integer", "number"):
        if f_min is not None:
            val = f_min - 1 if f_type == "integer" else round(f_min - 0.01, 2)
            results.append((val, f"最小值-1={val}(非法)"))
            results.append((f_min, f"最小值={f_min}(合法)"))
        if f_max is not None:
            results.append((f_max, f"最大值={f_max}(合法)"))
            val = f_max + 1 if f_type == "integer" else round(f_max + 0.01, 2)
            results.append((val, f"最大值+1={val}(非法)"))

    if f_enum:
        # 枚举外的值
        if f_type == "string":
            non_enum = "NON_ENUM_VALUE"
        else:
            non_enum = max(f_enum) + 1 if f_enum else 0
        results.append((non_enum, f"非枚举值={non_enum}(非法)"))

    return results


def _generate_spec_negative_for_field(field: dict) -> List[Tuple[Any, str]]:
    """为规范模式字段生成异常值"""
    results = []
    f_type = field.get("type", "string")
    f_name = field.get("name", "")
    f_required = field.get("required", False)
    f_format = field.get("format", "")
    f_enum = field.get("enum")

    # 空值
    if f_required:
        if f_type == "string":
            results.append(("", "空字符串"))
        results.append((None, "null值"))

    # 类型不匹配
    if f_type == "string":
        results.append((12345, "传整数"))
    elif f_type in ("integer", "number"):
        results.append(("not_a_number", "传字符串"))

    # 格式错误
    if f_format == "email" or "email" in f_name.lower():
        results.extend([(e, d) for e, d in _generate_invalid_email()])
    elif f_format == "uri" or f_format == "url" or "url" in f_name.lower():
        results.extend([(e, d) for e, d in _generate_invalid_url()])

    # 枚举外的值
    if f_enum:
        if f_type == "string":
            non_enum = "INVALID_ENUM"
        else:
            non_enum = max(f_enum) + 1 if f_enum else 0
        results.append((non_enum, "非枚举值"))

    # 超长/超大
    if f_type == "string":
        results.append(("a" * 1000, "超长字符串1000字符"))
    elif f_type in ("integer", "number"):
        results.append((999999999999, "超大数值"))

    # 安全注入（string 类型）
    if f_type == "string":
        results.append(("' OR '1'='1", "SQL注入"))
        results.append(("<script>alert(1)</script>", "XSS注入"))

    return results


def _write_spec_csv(all_results: dict, fields: list, output_dir: str, spec_name: str):
    """写入规范模式的 CSV 文件"""
    field_names = [f.get("name", "") for f in fields]

    for dim, rows in all_results.items():
        filename = f"spec_{spec_name}_{dim}.csv"
        filepath = os.path.join(output_dir, filename)

        with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            # 表头
            header = field_names[:]
            if dim in ("boundary", "negative"):
                header.append(f"_{dim}_field")
                header.append(f"_{dim}_desc")
            writer.writerow(header)

            for row in rows:
                csv_row = [row.get(fn, "") for fn in field_names]
                if dim in ("boundary", "negative"):
                    csv_row.append(row.get(f"_{dim}_field", ""))
                    csv_row.append(row.get(f"_{dim}_desc", ""))
                writer.writerow(csv_row)

        print(f"  输出文件: {filepath}")


def _write_spec_json(all_results: dict, fields: list, output_dir: str, spec_name: str):
    """写入规范模式的 JSON 文件"""
    for dim, rows in all_results.items():
        filename = f"spec_{spec_name}_{dim}.json"
        filepath = os.path.join(output_dir, filename)
        _write_json({"dimension": dim, "data": rows}, filepath)
        print(f"  输出文件: {filepath}")


def _write_spec_yaml(all_results: dict, fields: list, output_dir: str, spec_name: str):
    """写入规范模式的 YAML 文件"""
    for dim, rows in all_results.items():
        filename = f"spec_{spec_name}_{dim}.yaml"
        filepath = os.path.join(output_dir, filename)
        _write_yaml({"dimension": dim, "data": rows}, filepath)
        print(f"  输出文件: {filepath}")


def _write_spec_excel(all_results: dict, fields: list, output_dir: str, spec_name: str):
    """写入规范模式的 Excel 文件"""
    try:
        import openpyxl
    except ImportError:
        print("警告: openpyxl 未安装，回退到 JSON 输出", file=sys.stderr)
        _write_spec_json(all_results, fields, output_dir, spec_name)
        return

    field_names = [f.get("name", "") for f in fields]
    filename = f"spec_{spec_name}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()

    for dim, rows in all_results.items():
        ws = wb.create_sheet(title=dim)
        header = field_names[:]
        if dim in ("boundary", "negative"):
            header.append(f"变异字段")
            header.append(f"变异说明")
        ws.append(header)

        for row in rows:
            csv_row = [row.get(fn, "") for fn in field_names]
            if dim in ("boundary", "negative"):
                csv_row.append(row.get(f"_{dim}_field", ""))
                csv_row.append(row.get(f"_{dim}_desc", ""))
            ws.append(csv_row)

    # 删除默认 sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(filepath)
    print(f"  输出文件: {filepath}")


# ============================================================================
# 工具函数
# ============================================================================

def _generate_random_string(length: int, charset: str = string.ascii_lowercase + string.digits) -> str:
    """生成指定长度的随机字符串"""
    return ''.join(random.choice(charset) for _ in range(length))


def _extract_length_from_pattern(pattern: str) -> Tuple[Optional[int], Optional[int]]:
    """从正则表达式中提取隐含的长度约束"""
    if not pattern:
        return None, None

    inferred_min = None
    inferred_max = None

    all_quantifiers = re.findall(r'\{(\d+)(?:,\s*(\d*))?\}', pattern)

    for min_str, max_str in all_quantifiers:
        min_val = int(min_str)
        if inferred_min is None or min_val > inferred_min:
            inferred_min = min_val

        if max_str:
            max_val = int(max_str)
            if inferred_max is None or max_val < inferred_max:
                inferred_max = max_val
        elif ',' not in pattern[pattern.find('{' + min_str):pattern.find('}' , pattern.find('{' + min_str)) + 1]:
            inferred_max = min_val

    return inferred_min, inferred_max


def _generate_valid_value_for_pattern(pattern: str, length: int = 10) -> str:
    """根据正则表达式生成合法值（支持基础语法）"""
    common_patterns = {
        r"^(?=.*[A-Za-z])(?=.*\d).{8,}$": "Test1234",
        r"^[a-zA-Z0-9_]+$": "test_user01",
        r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$": "test@example.com",
        r"^1[3-9]\d{9}$": "13800001234",
        r"^https?://.+$": "https://example.com",
        r"^\d{4}-\d{2}-\d{2}$": "2026-01-15",
        r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}": "2026-01-15T10:30:00+08:00",
        r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$":
            "550e8400-e29b-41d4-a716-446655440000",
    }

    for pat, val in common_patterns.items():
        if re.search(pat.replace('^', '').replace('$', ''), pattern.replace('^', '').replace('$', '')):
            return val

    if '[' in pattern:
        char_class = re.findall(r'\[([^\]]+)\]', pattern)
        if char_class:
            chars = char_class[0].replace('a-z', 'a').replace('A-Z', 'A').replace('0-9', '0')
            return ''.join(random.choice(chars) for _ in range(length))

    return _generate_random_string(length)


def _get_type_default(param: dict) -> Any:
    """获取参数类型的默认合法值"""
    p_type = param.get("type", "string")
    p_format = param.get("format", "")
    p_pattern = param.get("pattern", "")
    p_name = param.get("name", "").lower()

    if "default" in param:
        return param["default"]
    if "example" in param:
        return param["example"]
    if "enum" in param and param["enum"]:
        return param["enum"][0]
    if p_pattern:
        return _generate_valid_value_for_pattern(p_pattern)

    format_defaults = {
        "email": "test@example.com",
        "uri": "https://example.com",
        "url": "https://example.com",
        "uuid": "550e8400-e29b-41d4-a716-446655440000",
        "date-time": "2026-01-15T10:30:00+08:00",
        "date": "2026-01-15",
        "password": "Test@1234",
        "int32": 1,
        "int64": 1,
    }
    if p_format in format_defaults:
        return format_defaults[p_format]

    name_defaults = {
        "username": "testuser01",
        "password": "Test@1234",
        "email": "test@example.com",
        "phone": "13800001234",
        "mobile": "13800001234",
        "url": "https://example.com",
        "avatar": "https://example.com/avatar.jpg",
        "nickname": "测试用户",
        "name": "测试",
        "title": "测试标题",
        "description": "测试描述",
        "remark": "测试备注",
        "address": "北京市朝阳区测试街道1号",
        "province": "北京市",
        "city": "北京市",
        "district": "朝阳区",
        "detail": "测试街道1号",
        "page": 1,
        "size": 20,
        "pagenum": 1,
        "pagesize": 20,
        "current": 1,
        "keyword": "测试",
        "status": 1,
        "selected": 1,
        "quantity": 1,
        "amount": 99.99,
        "price": 99.99,
        "stock": 100,
        "sort": 0,
        "gender": 1,
        "role": 0,
        "type": 1,
        "paymenttype": 1,
    }
    for key, val in name_defaults.items():
        if key in p_name:
            return val

    type_defaults = {
        "string": "test_value",
        "integer": 1,
        "number": 1.0,
        "boolean": True,
        "array": [],
        "object": {},
        "file": "(binary_file)",
    }
    return type_defaults.get(p_type, "test_value")


def _generate_string_by_length(length: int, pattern: str = None) -> str:
    """生成指定长度的字符串"""
    if length <= 0:
        return ""

    if pattern:
        base = _generate_valid_value_for_pattern(pattern, max(length, 8))
        if len(base) >= length:
            return base[:length]
        return base + "a" * (length - len(base))

    if length <= 5:
        return _generate_random_string(length, string.ascii_lowercase)
    elif length <= 20:
        return _generate_random_string(length, string.ascii_lowercase + string.digits)
    else:
        prefix = _generate_random_string(10, string.ascii_lowercase + string.digits)
        padding = "a" * (length - 10)
        return prefix + padding


def _identify_dependency(param_name: str, param_location: str, api_id: str = "") -> Optional[str]:
    """识别参数是否为依赖变量，返回变量名或 None"""
    name_lower = param_name.lower()

    if param_location == "header":
        if name_lower in ("authorization", "x-auth-token", "token"):
            return "${TOKEN}"

    for pattern, var_name, _ in ID_PARAM_PATTERNS:
        if re.match(pattern, param_name):
            return f"${{{var_name}}}"

    if param_location == "path" and ("id" in name_lower):
        path_part = api_id.split("_", 1)[-1] if "_" in api_id else ""
        resource_map = {
            "user": "USER_ID", "order": "ORDER_ID", "address": "ADDRESS_ID",
            "product": "PRODUCT_ID", "cart": "CART_ID", "banner": "BANNER_ID",
            "category": "CATEGORY_ID",
        }
        for key, var in resource_map.items():
            if key in path_part:
                return f"${{{var}}}"
        return "${RESOURCE_ID}"

    return None


# ============================================================================
# 传统模式 - 正向数据生成
# ============================================================================

def generate_positive_data(api_def: dict, case_counter: dict) -> List[dict]:
    """生成正向合法数据"""
    cases = []
    params = api_def.get("parameters", {})
    all_params = _collect_all_params(params)

    if not all_params:
        cases.append(_build_test_case(
            case_counter, "合法请求", "positive", "P0",
            _build_empty_params(), api_def
        ))
        return cases

    required_params = {k: v for k, v in all_params.items() if v.get("required")}
    optional_params = {k: v for k, v in all_params.items() if not v.get("required")}

    full_data = _generate_legal_values(all_params)
    cases.append(_build_test_case(
        case_counter, "所有参数合法-完整参数", "positive", "P0",
        full_data, api_def
    ))

    if optional_params:
        required_data = _generate_legal_values(required_params)
        cases.append(_build_test_case(
            case_counter, "仅必填参数合法", "positive", "P1",
            required_data, api_def
        ))

    for param_key, param_def in all_params.items():
        if "enum" in param_def and len(param_def["enum"]) > 1:
            for enum_val in param_def["enum"][1:]:
                enum_data = copy.deepcopy(full_data)
                param_location = param_def.get("in", "body")
                param_name = param_def.get("name", "")
                _set_param_value(enum_data, param_location, param_name, enum_val)
                cases.append(_build_test_case(
                    case_counter,
                    f"枚举遍历-{param_name}={enum_val}",
                    "positive", "P2",
                    enum_data, api_def
                ))

    return cases


# ============================================================================
# 传统模式 - 边界值数据生成
# ============================================================================

def generate_boundary_data(api_def: dict, case_counter: dict) -> List[dict]:
    """生成边界值数据"""
    cases = []
    params = api_def.get("parameters", {})
    all_params = _collect_all_params(params)

    for param_key, param_def in all_params.items():
        p_type = param_def.get("type", "string")

        if p_type == "string":
            cases.extend(_generate_string_boundary(
                case_counter, param_def, all_params, api_def
            ))
        elif p_type in ("integer", "number"):
            cases.extend(_generate_number_boundary(
                case_counter, param_def, all_params, api_def
            ))
        elif p_type == "array":
            cases.extend(_generate_array_boundary(
                case_counter, param_def, all_params, api_def
            ))

    return cases


def _generate_string_boundary(case_counter: dict, param_def: dict,
                               all_params: dict, api_def: dict) -> List[dict]:
    """生成字符串参数的边界值"""
    cases = []
    p_name = param_def.get("name", "")
    p_location = param_def.get("in", "body")
    min_len = param_def.get("minLength", DEFAULT_MIN_LENGTH)
    max_len = param_def.get("maxLength", DEFAULT_MAX_LENGTH)
    pattern = param_def.get("pattern", "")

    inferred_min, inferred_max = _extract_length_from_pattern(pattern)
    if inferred_min is not None and (min_len == DEFAULT_MIN_LENGTH or inferred_min > min_len):
        min_len = inferred_min
    if inferred_max is not None and (max_len == DEFAULT_MAX_LENGTH or inferred_max < max_len):
        max_len = inferred_max

    if min_len == DEFAULT_MIN_LENGTH and max_len == DEFAULT_MAX_LENGTH and not pattern:
        return cases

    legal_values = _generate_legal_values(all_params)

    if min_len > 0:
        below_min = copy.deepcopy(legal_values)
        _set_param_value(below_min, p_location, p_name,
                        _generate_string_by_length(max(0, min_len - 1), pattern or None))
        cases.append(_build_test_case(
            case_counter, f"{p_name}最小长度边界-{min_len - 1}字符(非法)",
            "boundary", "P1", below_min, api_def,
            expected_status=400
        ))

        at_min = copy.deepcopy(legal_values)
        _set_param_value(at_min, p_location, p_name,
                        _generate_string_by_length(min_len, pattern or None))
        cases.append(_build_test_case(
            case_counter, f"{p_name}最小长度边界-{min_len}字符(合法)",
            "boundary", "P0", at_min, api_def
        ))

    if max_len < 10000:
        at_max = copy.deepcopy(legal_values)
        _set_param_value(at_max, p_location, p_name,
                        _generate_string_by_length(max_len, pattern or None))
        cases.append(_build_test_case(
            case_counter, f"{p_name}最大长度边界-{max_len}字符(合法)",
            "boundary", "P0", at_max, api_def
        ))

        above_max = copy.deepcopy(legal_values)
        _set_param_value(above_max, p_location, p_name,
                        _generate_string_by_length(max_len + 1, pattern or None))
        cases.append(_build_test_case(
            case_counter, f"{p_name}最大长度边界-{max_len + 1}字符(非法)",
            "boundary", "P1", above_max, api_def,
            expected_status=400
        ))

    return cases


def _generate_number_boundary(case_counter: dict, param_def: dict,
                               all_params: dict, api_def: dict) -> List[dict]:
    """生成数值参数的边界值"""
    cases = []
    p_name = param_def.get("name", "")
    p_location = param_def.get("in", "body")
    minimum = param_def.get("minimum")
    maximum = param_def.get("maximum")

    if minimum is None and maximum is None:
        return cases

    legal_values = _generate_legal_values(all_params)

    if minimum is not None:
        below_min = copy.deepcopy(legal_values)
        _set_param_value(below_min, p_location, p_name, minimum - 1)
        cases.append(_build_test_case(
            case_counter, f"{p_name}最小值边界-{minimum - 1}(非法)",
            "boundary", "P1", below_min, api_def,
            expected_status=400
        ))

        at_min = copy.deepcopy(legal_values)
        _set_param_value(at_min, p_location, p_name, minimum)
        cases.append(_build_test_case(
            case_counter, f"{p_name}最小值边界-{minimum}(合法)",
            "boundary", "P0", at_min, api_def
        ))

    if maximum is not None:
        at_max = copy.deepcopy(legal_values)
        _set_param_value(at_max, p_location, p_name, maximum)
        cases.append(_build_test_case(
            case_counter, f"{p_name}最大值边界-{maximum}(合法)",
            "boundary", "P0", at_max, api_def
        ))

        above_max = copy.deepcopy(legal_values)
        _set_param_value(above_max, p_location, p_name, maximum + 1)
        cases.append(_build_test_case(
            case_counter, f"{p_name}最大值边界-{maximum + 1}(非法)",
            "boundary", "P1", above_max, api_def,
            expected_status=400
        ))

    if minimum is not None and minimum > 0:
        zero_data = copy.deepcopy(legal_values)
        _set_param_value(zero_data, p_location, p_name, 0)
        cases.append(_build_test_case(
            case_counter, f"{p_name}零值测试",
            "boundary", "P1", zero_data, api_def,
            expected_status=400
        ))

    if minimum is not None and minimum >= 0:
        neg_data = copy.deepcopy(legal_values)
        _set_param_value(neg_data, p_location, p_name, -1)
        cases.append(_build_test_case(
            case_counter, f"{p_name}负数测试",
            "boundary", "P2", neg_data, api_def,
            expected_status=400
        ))

    return cases


def _generate_array_boundary(case_counter: dict, param_def: dict,
                              all_params: dict, api_def: dict) -> List[dict]:
    """生成数组参数的边界值"""
    cases = []
    p_name = param_def.get("name", "")
    p_location = param_def.get("in", "body")
    min_items = param_def.get("minItems", DEFAULT_MIN_ITEMS)
    max_items = param_def.get("maxItems", DEFAULT_MAX_ITEMS)

    if min_items == DEFAULT_MIN_ITEMS and max_items == DEFAULT_MAX_ITEMS:
        return cases

    legal_values = _generate_legal_values(all_params)
    item_type = param_def.get("items", {}).get("type", "string")

    def _make_array(count):
        if item_type == "integer":
            return list(range(1, count + 1))
        elif item_type == "string":
            return [f"item_{i}" for i in range(count)]
        else:
            return [{} for _ in range(count)]

    empty_arr = copy.deepcopy(legal_values)
    _set_param_value(empty_arr, p_location, p_name, [])
    cases.append(_build_test_case(
        case_counter, f"{p_name}空数组",
        "boundary", "P1" if min_items > 0 else "P2", empty_arr, api_def,
        expected_status=400 if min_items > 0 else None
    ))

    if min_items > 0:
        min_arr = copy.deepcopy(legal_values)
        _set_param_value(min_arr, p_location, p_name, _make_array(min_items))
        cases.append(_build_test_case(
            case_counter, f"{p_name}最小元素数-{min_items}(合法)",
            "boundary", "P0", min_arr, api_def
        ))

    if max_items < 100:
        max_arr = copy.deepcopy(legal_values)
        _set_param_value(max_arr, p_location, p_name, _make_array(max_items))
        cases.append(_build_test_case(
            case_counter, f"{p_name}最大元素数-{max_items}(合法)",
            "boundary", "P0", max_arr, api_def
        ))

    return cases


# ============================================================================
# 传统模式 - 异常数据生成
# ============================================================================

def generate_negative_data(api_def: dict, case_counter: dict) -> List[dict]:
    """生成异常非法数据"""
    cases = []
    params = api_def.get("parameters", {})
    all_params = _collect_all_params(params)

    for param_key, param_def in all_params.items():
        p_type = param_def.get("type", "string")
        p_name = param_def.get("name", "")
        p_location = param_def.get("in", "body")
        p_required = param_def.get("required", False)

        legal_values = _generate_legal_values(all_params)

        if p_required:
            empty_data = copy.deepcopy(legal_values)
            _set_param_value(empty_data, p_location, p_name, "")
            cases.append(_build_test_case(
                case_counter, f"{p_name}为空字符串",
                "negative", "P0", empty_data, api_def,
                expected_status=400
            ))

            missing_data = copy.deepcopy(legal_values)
            _remove_param(missing_data, p_location, p_name)
            cases.append(_build_test_case(
                case_counter, f"{p_name}缺失",
                "negative", "P0", missing_data, api_def,
                expected_status=400
            ))

        null_data = copy.deepcopy(legal_values)
        _set_param_value(null_data, p_location, p_name, None)
        cases.append(_build_test_case(
            case_counter, f"{p_name}为null",
            "negative", "P1", null_data, api_def,
            expected_status=400
        ))

        if p_type == "string":
            type_data = copy.deepcopy(legal_values)
            _set_param_value(type_data, p_location, p_name, 12345)
            cases.append(_build_test_case(
                case_counter, f"{p_name}类型不匹配-传整数",
                "negative", "P2", type_data, api_def,
                expected_status=400
            ))
        elif p_type in ("integer", "number"):
            type_data = copy.deepcopy(legal_values)
            _set_param_value(type_data, p_location, p_name, "not_a_number")
            cases.append(_build_test_case(
                case_counter, f"{p_name}类型不匹配-传字符串",
                "negative", "P1", type_data, api_def,
                expected_status=400
            ))

        p_format = param_def.get("format", "")
        format_errors = _get_format_error_values(p_format, p_name)
        for err_val, err_desc in format_errors:
            fmt_data = copy.deepcopy(legal_values)
            _set_param_value(fmt_data, p_location, p_name, err_val)
            cases.append(_build_test_case(
                case_counter, f"{p_name}格式错误-{err_desc}",
                "negative", "P1", fmt_data, api_def,
                expected_status=400
            ))

        if "enum" in param_def:
            enum_data = copy.deepcopy(legal_values)
            non_enum_val = _get_non_enum_value(param_def["enum"], p_type)
            _set_param_value(enum_data, p_location, p_name, non_enum_val)
            cases.append(_build_test_case(
                case_counter, f"{p_name}非枚举值-{non_enum_val}",
                "negative", "P0", enum_data, api_def,
                expected_status=400
            ))

        if p_type == "string":
            long_data = copy.deepcopy(legal_values)
            _set_param_value(long_data, p_location, p_name, "a" * 1000)
            cases.append(_build_test_case(
                case_counter, f"{p_name}超长字符串-1000字符",
                "negative", "P2", long_data, api_def,
                expected_status=400
            ))

        if p_type in ("integer", "number"):
            huge_data = copy.deepcopy(legal_values)
            _set_param_value(huge_data, p_location, p_name, 999999999999)
            cases.append(_build_test_case(
                case_counter, f"{p_name}超大数值",
                "negative", "P2", huge_data, api_def,
                expected_status=400
            ))

    return cases


def _get_format_error_values(fmt: str, name: str) -> List[Tuple[str, str]]:
    """获取格式约束的错误值"""
    errors = []
    if fmt == "email" or "email" in name.lower():
        errors.extend([
            ("test@", "缺域名"),
            ("@example.com", "缺用户名"),
            ("test@example", "缺顶级域名"),
            ("test", "无@符号"),
        ])
    elif fmt == "uri" or fmt == "url" or "url" in name.lower():
        errors.extend([
            ("example.com", "缺协议"),
            ("ftp://example.com", "非HTTP协议"),
        ])
    elif fmt == "date-time" or "date" in name.lower():
        errors.extend([
            ("2026/01/15", "日期格式错误-斜杠"),
            ("not-a-date", "非法日期"),
            ("2026-13-01", "月份超范围"),
        ])
    elif fmt == "uuid":
        errors.extend([
            ("not-a-uuid", "非法UUID"),
            ("550e8400-e29b-41d4", "UUID不完整"),
        ])
    elif fmt == "phone" or "phone" in name.lower() or "mobile" in name.lower():
        errors.extend([
            ("1234567", "电话号码过短"),
            ("abc12345678", "含非数字字符"),
            ("123456789012345", "电话号码过长"),
        ])
    return errors


def _get_non_enum_value(enum_list: list, p_type: str) -> Any:
    """获取不在枚举列表中的值"""
    if p_type == "string":
        for candidate in ["INVALID_ENUM", "unknown", "other", "X"]:
            if candidate not in enum_list:
                return candidate
        return "NON_ENUM_VALUE_" + _generate_random_string(4)
    elif p_type in ("integer", "number"):
        for i in range(-1, 1000):
            if i not in enum_list:
                return i
        return max(enum_list) + 1 if enum_list else 0
    else:
        return "INVALID"


# ============================================================================
# 传统模式 - 安全数据生成
# ============================================================================

def generate_security_data(api_def: dict, case_counter: dict) -> List[dict]:
    """生成安全与幂等数据"""
    cases = []
    params = api_def.get("parameters", {})
    all_params = _collect_all_params(params)

    for param_key, param_def in all_params.items():
        p_type = param_def.get("type", "string")
        p_name = param_def.get("name", "")
        p_location = param_def.get("in", "body")

        if p_type not in ("string", "integer", "number"):
            continue

        if p_location == "header" and p_name.lower() in ("authorization",):
            continue

        legal_values = _generate_legal_values(all_params)

        if p_type == "string":
            for payload in SQL_INJECTION_PAYLOADS:
                sec_data = copy.deepcopy(legal_values)
                _set_param_value(sec_data, p_location, p_name, payload["payload"])
                cases.append(_build_test_case(
                    case_counter, f"SQL注入-{p_name}-{payload['id']}",
                    "security", "P1", sec_data, api_def,
                    expected_status=400,
                    security_info=payload
                ))

            for payload in XSS_PAYLOADS:
                sec_data = copy.deepcopy(legal_values)
                _set_param_value(sec_data, p_location, p_name, payload["payload"])
                cases.append(_build_test_case(
                    case_counter, f"XSS攻击-{p_name}-{payload['id']}",
                    "security", "P1", sec_data, api_def,
                    expected_status=400,
                    security_info=payload
                ))

            for payload in PATH_TRAVERSAL_PAYLOADS:
                sec_data = copy.deepcopy(legal_values)
                _set_param_value(sec_data, p_location, p_name, payload["payload"])
                cases.append(_build_test_case(
                    case_counter, f"路径穿越-{p_name}-{payload['id']}",
                    "security", "P2", sec_data, api_def,
                    expected_status=400,
                    security_info=payload
                ))

            for payload in COMMAND_INJECTION_PAYLOADS:
                sec_data = copy.deepcopy(legal_values)
                _set_param_value(sec_data, p_location, p_name, payload["payload"])
                cases.append(_build_test_case(
                    case_counter, f"命令注入-{p_name}-{payload['id']}",
                    "security", "P2", sec_data, api_def,
                    expected_status=400,
                    security_info=payload
                ))

    if api_def.get("method") == "POST":
        legal_values = _generate_legal_values(all_params)
        cases.append(_build_test_case(
            case_counter, "重复提交-同请求发送2次",
            "security", "P2", legal_values, api_def,
            idempotent=True
        ))

    return cases


# ============================================================================
# 传统模式 - 参数集合与值设置辅助函数
# ============================================================================

def _collect_all_params(params: dict) -> dict:
    """收集所有参数，返回 {param_key: param_def} 字典"""
    result = {}
    for param_type in ["path_params", "query_params", "header_params", "body_params"]:
        for param in params.get(param_type, []):
            p_name = param.get("name", "")
            p_in = param.get("in", param_type.replace("_params", ""))
            key = f"{p_in}:{p_name}"
            param = _infer_required(param, p_in)
            result[key] = param
            for prop in param.get("properties", []):
                prop_name = prop.get("name", "")
                prop_key = f"body:{p_name}.{prop_name}"
                prop = _infer_required(prop, "body")
                result[prop_key] = prop
                prop["in"] = "body"
    return result


def _infer_required(param: dict, p_in: str) -> dict:
    """推断参数必填性"""
    param = copy.deepcopy(param)
    if param.get("required"):
        return param
    if p_in == "path":
        param["required"] = True
        param["required_inferred"] = True
        return param
    if param.get("pattern"):
        param["required"] = True
        param["required_inferred"] = True
        return param
    required_names = {
        "username", "password", "email", "phone", "mobile",
        "name", "title", "content", "description",
        "orderId", "productId", "addressId", "userId",
        "quantity", "price", "amount", "total",
        "captchaKey", "captchaCode",
        "oldPassword", "newPassword",
    }
    p_name = param.get("name", "")
    if p_name in required_names:
        param["required"] = True
        param["required_inferred"] = True
        return param
    if re.search(r'(Id|Code|Key)$', p_name):
        param["required"] = True
        param["required_inferred"] = True
        return param
    return param


def _generate_legal_values(all_params: dict) -> dict:
    """为所有参数生成合法值"""
    result = {
        "path_params": {},
        "query_params": {},
        "header_params": {},
        "body_params": {},
    }

    for param_key, param_def in all_params.items():
        p_name = param_def.get("name", "")
        p_in = param_def.get("in", "body")

        dep = _identify_dependency(p_name, p_in)
        value = dep if dep else _get_type_default(param_def)

        target = f"{p_in}_params"
        if target in result:
            result[target][p_name] = value

    return result


def _build_empty_params() -> dict:
    """构建空参数结构"""
    return {
        "path_params": {},
        "query_params": {},
        "header_params": {},
        "body_params": {},
    }


def _set_param_value(params_data: dict, location: str, name: str, value: Any):
    """设置参数值"""
    target = f"{location}_params"
    if target in params_data:
        params_data[target][name] = value


def _remove_param(params_data: dict, location: str, name: str):
    """移除参数"""
    target = f"{location}_params"
    if target in params_data and name in params_data[target]:
        del params_data[target][name]


def _build_test_case(case_counter: dict, name: str, category: str,
                     priority: str, parameters: dict, api_def: dict,
                     expected_status: int = None, security_info: dict = None,
                     idempotent: bool = False) -> dict:
    """构建测试用例数据"""
    category_prefix = {
        "positive": "POS",
        "boundary": "BND",
        "negative": "NEG",
        "security": "SEC",
    }

    if category not in case_counter:
        case_counter[category] = 0
    case_counter[category] += 1

    case_id = f"{category_prefix.get(category, 'TC')}_{case_counter[category]:03d}"

    expected = {"status_code": expected_status or 200}
    if expected_status and expected_status >= 400:
        expected["assertions"] = [
            {"field": "code", "operator": "neq", "value": 200}
        ]
    elif category == "positive":
        expected["assertions"] = [
            {"field": "code", "operator": "eq", "value": 200}
        ]

    test_case = {
        "case_id": case_id,
        "name": name,
        "category": category,
        "priority": priority,
        "parameters": parameters,
        "expected": expected,
    }

    if security_info:
        test_case["security_type"] = security_info.get("id", "").split("-")[0].lower() + "_injection"
        test_case["attack_vector"] = security_info.get("id", "")

    if idempotent:
        test_case["idempotent"] = True
        test_case["repeat_count"] = 2

    return test_case


# ============================================================================
# 传统模式 - 依赖分析
# ============================================================================

def analyze_dependencies(apis: list) -> dict:
    """分析所有接口的依赖关系"""
    dependencies = {
        "global_dependencies": [],
        "dependency_chains": [],
        "resource_map": {},
    }

    has_auth = False
    for api in apis:
        for header_param in api.get("parameters", {}).get("header_params", []):
            if header_param.get("name", "").lower() in ("authorization",):
                has_auth = True
                break
        if has_auth:
            break

    login_api = None
    for api in apis:
        if "login" in api.get("api_id", "").lower():
            login_api = api
            break

    if has_auth and login_api:
        dependencies["global_dependencies"].append({
            "variable": "TOKEN",
            "source_api": login_api["api_id"],
            "extract_path": "data.token",
            "description": "登录获取 Token",
            "source_params": _generate_legal_values(
                _collect_all_params(login_api.get("parameters", {}))
            ),
        })

    resource_map = {}
    for api in apis:
        path = api.get("path", "")
        method = api.get("method", "")
        path_parts = [p for p in path.split("/") if p and not p.startswith("{")]
        if not path_parts:
            continue
        resource_name = path_parts[-1] if not path.endswith("}") else path_parts[-2] if len(path_parts) >= 2 else None
        if not resource_name:
            continue
        if resource_name not in resource_map:
            resource_map[resource_name] = {"create_api": None, "operation_apis": []}
        if method == "POST" and not path.endswith("}"):
            resource_map[resource_name]["create_api"] = api
        elif method in ("GET", "PUT", "DELETE") and path.endswith("}"):
            resource_map[resource_name]["operation_apis"].append(api)

    dependencies["resource_map"] = {
        k: {
            "create_api": v["create_api"]["api_id"] if v["create_api"] else None,
            "operation_apis": [a["api_id"] for a in v["operation_apis"]],
        }
        for k, v in resource_map.items()
    }

    for resource_name, res_info in resource_map.items():
        create_api = res_info["create_api"]
        if not create_api:
            continue
        chain = {"name": f"{resource_name}流程", "steps": []}
        chain["steps"].append({
            "step": 1,
            "api_id": create_api["api_id"],
            "description": f"创建{resource_name}",
            "requires": ["TOKEN"] if has_auth else [],
            "extract": {f"{resource_name.upper()}_ID": "data.id"},
        })
        for i, op_api in enumerate(res_info["operation_apis"], 2):
            chain["steps"].append({
                "step": i,
                "api_id": op_api["api_id"],
                "description": op_api.get("name", ""),
                "requires": ["TOKEN"] if has_auth else [],
            })
        if len(chain["steps"]) > 1:
            dependencies["dependency_chains"].append(chain)

    return dependencies


# ============================================================================
# 传统模式 - 生成与输出
# ============================================================================

def generate_test_data(input_path: str, output_dir: str, output_format: str = "yaml",
                       module_filter: str = None, api_filter: str = None,
                       config_path: str = None, dimensions: list = None) -> dict:
    """传统模式主生成函数"""
    with open(input_path, 'r', encoding='utf-8') as f:
        definition = json.load(f)

    apis = definition.get("apis", [])
    meta = definition.get("meta", {})
    global_rules = definition.get("global_rules", {})

    custom_config = {}
    if config_path and os.path.exists(config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            if config_path.endswith('.json'):
                custom_config = json.load(f)
            else:
                import yaml
                custom_config = yaml.safe_load(f)

    if not dimensions:
        dimensions = ["positive", "boundary", "negative", "security"]

    filtered_apis = apis
    if module_filter:
        filtered_apis = [a for a in filtered_apis if a.get("module") == module_filter]
    if api_filter:
        filtered_apis = [a for a in filtered_apis if a.get("api_id") == api_filter]

    os.makedirs(output_dir, exist_ok=True)

    stats = {
        "total_apis": len(filtered_apis),
        "total_cases": 0,
        "by_category": {},
        "by_module": {},
    }

    module_groups = {}
    for api in filtered_apis:
        module = api.get("module", "未分类")
        if module not in module_groups:
            module_groups[module] = []
        module_groups[module].append(api)

    dep_analysis = analyze_dependencies(filtered_apis)

    all_api_results = {}
    for module_name, module_apis in module_groups.items():
        module_dir_name = _to_dir_name(module_name)
        module_dir = os.path.join(output_dir, module_dir_name)
        os.makedirs(module_dir, exist_ok=True)

        for api in module_apis:
            case_counter = {}
            cases = []

            if "positive" in dimensions:
                cases.extend(generate_positive_data(api, case_counter))
            if "boundary" in dimensions:
                cases.extend(generate_boundary_data(api, case_counter))
            if "negative" in dimensions:
                cases.extend(generate_negative_data(api, case_counter))
            if "security" in dimensions:
                cases.extend(generate_security_data(api, case_counter))

            api_result = {
                "api_id": api.get("api_id", ""),
                "name": api.get("name", ""),
                "module": module_name,
                "description": api.get("description", ""),
                "method": api.get("method", ""),
                "path": api.get("path", ""),
                "test_cases": cases,
            }

            all_api_results[api["api_id"]] = api_result
            _write_api_data(api_result, module_dir, output_format)

            stats["total_cases"] += len(cases)
            for case in cases:
                cat = case.get("category", "unknown")
                stats["by_category"][cat] = stats["by_category"].get(cat, 0) + 1

            if module_name not in stats["by_module"]:
                stats["by_module"][module_name] = {"api_count": 0, "case_count": 0}
            stats["by_module"][module_name]["api_count"] += 1
            stats["by_module"][module_name]["case_count"] += len(cases)

    _write_dependencies(dep_analysis, output_dir, output_format)
    _write_config(definition, output_dir, output_format)

    summary = {
        "meta": {
            "version": VERSION,
            "generated_at": datetime.now(timezone(timedelta(hours=8))).isoformat(),
            "source_file": os.path.basename(input_path),
        },
        "stats": stats,
        "dimensions": dimensions,
        "modules": list(module_groups.keys()),
    }
    _write_json(summary, os.path.join(output_dir, "summary.json"))

    return summary


def _to_dir_name(module_name: str) -> str:
    """将模块名称转为目录名"""
    name_map = {
        "认证管理": "auth",
        "用户管理": "user",
        "商品管理": "product",
        "订单管理": "order",
        "地址管理": "address",
        "购物车管理": "cart",
        "验证码管理": "captcha",
        "管理后台": "admin",
        "商品搜索": "search",
        "Banner轮播图": "banner",
    }
    return name_map.get(module_name, module_name.lower().replace(" ", "_"))


def _write_api_data(api_result: dict, module_dir: str, fmt: str):
    """写入单个接口的测试数据文件"""
    api_id = api_result["api_id"]
    safe_name = api_id.replace("/", "_").replace("{", "").replace("}", "").lower()
    safe_name = re.sub(r'_+', '_', safe_name).strip('_')
    file_path = os.path.join(module_dir, f"{safe_name}")

    if fmt == "yaml":
        _write_yaml(api_result, f"{file_path}.yaml")
    elif fmt == "json":
        _write_json(api_result, f"{file_path}.json")
    elif fmt == "excel":
        _write_excel(api_result, f"{file_path}.xlsx")


def _write_dependencies(dep_analysis: dict, output_dir: str, fmt: str):
    """写入依赖配置文件"""
    file_path = os.path.join(output_dir, "_dependencies")
    if fmt == "yaml":
        _write_yaml(dep_analysis, f"{file_path}.yaml")
    else:
        _write_json(dep_analysis, f"{file_path}.json")


def _write_config(definition: dict, output_dir: str, fmt: str):
    """写入全局配置文件"""
    config = {
        "base_url": "http://localhost:8080",
        "auth": {
            "type": "Bearer Token",
            "header": "Authorization",
            "token_variable": "TOKEN",
            "default_user": {
                "username": "admin",
                "password": "Admin1234",
            },
        },
        "test_accounts": [
            {"username": "admin", "password": "Admin1234", "role": "admin", "description": "管理员账号"},
            {"username": "testuser01", "password": "Test1234", "role": "user", "description": "普通用户账号"},
        ],
        "response_format": {
            "success": {"code": 200, "data_path": "data"},
            "error": {"code_path": "code", "message_path": "message"},
        },
    }

    file_path = os.path.join(output_dir, "_config")
    if fmt == "yaml":
        _write_yaml(config, f"{file_path}.yaml")
    else:
        _write_json(config, f"{file_path}.json")


def _write_yaml(data: dict, file_path: str):
    """写入 YAML 文件"""
    try:
        import yaml
    except ImportError:
        _write_json(data, file_path.replace(".yaml", ".json"))
        return

    with open(file_path, 'w', encoding='utf-8') as f:
        yaml.dump(data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)


def _write_json(data: dict, file_path: str):
    """写入 JSON 文件"""
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _write_excel(api_result: dict, file_path: str):
    """写入 Excel 文件"""
    try:
        import openpyxl
    except ImportError:
        _write_json(api_result, file_path.replace(".xlsx", ".json"))
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试数据"

    headers = ["用例编号", "用例名称", "分类", "优先级", "Path参数", "Query参数",
               "Header参数", "Body参数", "预期状态码", "安全类型", "攻击向量"]
    ws.append(headers)

    for case in api_result.get("test_cases", []):
        params = case.get("parameters", {})
        row = [
            case.get("case_id", ""),
            case.get("name", ""),
            case.get("category", ""),
            case.get("priority", ""),
            json.dumps(params.get("path_params", {}), ensure_ascii=False),
            json.dumps(params.get("query_params", {}), ensure_ascii=False),
            json.dumps(params.get("header_params", {}), ensure_ascii=False),
            json.dumps(params.get("body_params", {}), ensure_ascii=False),
            case.get("expected", {}).get("status_code", ""),
            case.get("security_type", ""),
            case.get("attack_vector", ""),
        ]
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    wb.save(file_path)


# ============================================================================
# 模式自动识别
# ============================================================================

def detect_mode(input_path: str = None, mode_arg: str = None, prompt_arg: str = None) -> str:
    """自动识别运行模式

    Returns:
        "api" | "spec" | "nl"
    """
    if mode_arg:
        return mode_arg

    if prompt_arg:
        return "nl"

    if input_path and os.path.exists(input_path):
        try:
            with open(input_path, 'r', encoding='utf-8') as f:
                data = json.load(f) if input_path.endswith('.json') else {}
            if "apis" in data:
                return "api"
            elif "fields" in data:
                return "spec"
        except Exception:
            pass

    # 默认
    if input_path:
        return "api"

    return "nl"


# ============================================================================
# CLI 入口
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="API Test Data Generator - 测试数据自动化构造引擎 v2.0",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 传统模式 - 基于接口定义生成测试数据
  python3 testdata_generator.py api_definitions.json --format yaml --output test_data/

  # 规范模式 - 基于自定义字段规则生成数据
  python3 testdata_generator.py --mode spec --spec rules.yaml --output spec_data/

  # 自然语言模式 - 基于自然语言描述生成数据
  python3 testdata_generator.py --mode nl --prompt "生成10条符合中国身份证规则的数据" --output nl_data/
  python3 testdata_generator.py --mode nl --prompt "生成5个合法手机号和2个非法手机号" --output nl_data/
  python3 testdata_generator.py --mode nl --prompt "生成随机用户名、邮箱、地址、公司名等测试数据" --output nl_data/
        """
    )
    parser.add_argument("input", nargs='?', default=None,
                       help="输入文件路径（api_definitions.json 或字段规则文件）")
    parser.add_argument("--mode", "-m", choices=["api", "spec", "nl"],
                       default=None,
                       help="运行模式：api(传统)/spec(规范)/nl(自然语言)，不指定则自动识别")
    parser.add_argument("--spec", default=None,
                       help="规范模式的字段规则文件路径")
    parser.add_argument("--prompt", "-p", default=None,
                       help="自然语言模式的描述文本")
    parser.add_argument("--format", choices=["yaml", "json", "excel", "csv"],
                       default=None,
                       help="输出格式 (传统模式默认yaml, 规范模式默认csv, 自然语言模式默认csv)")
    parser.add_argument("--output", "-o", default="./test_data",
                       help="输出目录 (默认: ./test_data)")
    parser.add_argument("--module", default=None,
                       help="仅生成指定模块的测试数据（传统模式）")
    parser.add_argument("--api", default=None,
                       help="仅生成指定接口的测试数据（传统模式，api_id）")
    parser.add_argument("--config", "-c", default=None,
                       help="自定义团队数据规范文件路径（传统模式）")
    parser.add_argument("--dimension", "-d", default=None,
                       help="生成维度，逗号分隔 (positive,boundary,negative,security)")

    args = parser.parse_args()

    # 识别模式
    mode = detect_mode(
        input_path=args.input,
        mode_arg=args.mode,
        prompt_arg=args.prompt
    )

    # 确定 prompt 来源
    prompt = args.prompt
    if not prompt and mode == "nl" and args.input:
        # 如果用户把自然语言描述作为 positional arg 传入
        if not os.path.exists(args.input):
            prompt = args.input
            args.input = None
            mode = "nl"

    # 确定默认输出格式
    if args.format is None:
        if mode == "api":
            output_format = "yaml"
        elif mode == "spec":
            output_format = "csv"
        else:
            output_format = "csv"
    else:
        output_format = args.format

    # 解析维度
    dimensions = None
    if args.dimension:
        dimensions = [d.strip() for d in args.dimension.split(",")]

    # 执行
    if mode == "nl":
        if not prompt:
            print("错误: 自然语言模式需要提供 --prompt 参数", file=sys.stderr)
            sys.exit(1)

        print(f"自然语言模式: {prompt}")
        summary = generate_nl_data(
            prompt=prompt,
            output_dir=args.output,
        )

        print(f"\n生成完成！")
        print(f"  解析字段: {', '.join(summary.get('fields', []))}")
        print(f"  数据行数: {summary.get('total_rows', 0)}")
        if 'stats' in summary:
            for ft, stat in summary['stats'].items():
                display = NL_FIELD_DISPLAY_NAMES.get(ft, ft)
                parts = []
                if stat.get('valid', 0) > 0:
                    parts.append(f"合法{stat['valid']}条")
                if stat.get('invalid', 0) > 0:
                    parts.append(f"非法{stat['invalid']}条")
                if stat.get('boundary', 0) > 0:
                    parts.append(f"边界{stat['boundary']}条")
                print(f"    {display}: {', '.join(parts)}")
        print(f"  输出目录: {args.output}")

    elif mode == "spec":
        spec_path = args.spec or args.input
        if not spec_path or not os.path.exists(spec_path):
            print("错误: 规范模式需要提供 --spec 参数或输入文件路径", file=sys.stderr)
            sys.exit(1)

        print(f"规范模式: {spec_path}")
        summary = generate_spec_data(
            spec_path=spec_path,
            output_dir=args.output,
            output_format=output_format,
            dimensions=dimensions,
        )

        print(f"\n生成完成！")
        print(f"  规则名称: {summary.get('spec_name', 'N/A')}")
        print(f"  字段数量: {summary.get('field_count', 0)}")
        print(f"  数据行数: {summary.get('stats', {}).get('total_rows', 0)}")
        if 'stats' in summary and 'by_dimension' in summary['stats']:
            print(f"  维度统计:")
            for dim, count in summary['stats']['by_dimension'].items():
                dim_names = {"positive": "正向数据", "boundary": "边界值", "negative": "异常数据", "security": "安全数据"}
                print(f"    {dim_names.get(dim, dim)}: {count}")
        print(f"  输出目录: {args.output}")
        print(f"  输出格式: {output_format}")

    else:  # mode == "api"
        if not args.input or not os.path.exists(args.input):
            print("错误: 传统模式需要提供 api_definitions.json 文件路径", file=sys.stderr)
            sys.exit(1)

        print(f"传统模式: {args.input}")
        summary = generate_test_data(
            input_path=args.input,
            output_dir=args.output,
            output_format=output_format,
            module_filter=args.module,
            api_filter=args.api,
            config_path=args.config,
            dimensions=dimensions,
        )

        print(f"\n生成完成！")
        print(f"  接口数量: {summary['stats']['total_apis']}")
        print(f"  测试数据总数: {summary['stats']['total_cases']}")
        print(f"  分类统计:")
        for cat, count in summary['stats']['by_category'].items():
            cat_names = {"positive": "正向数据", "boundary": "边界值", "negative": "异常数据", "security": "安全数据"}
            print(f"    {cat_names.get(cat, cat)}: {count}")
        print(f"  输出目录: {args.output}")
        print(f"  输出格式: {output_format}")


if __name__ == "__main__":
    main()
